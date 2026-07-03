#!/usr/bin/env python3
"""Sync de maestro: lee lo manual (Pedido) del maestro para arrastrarlo al
regenerarlo desde el crudo nuevo, y compara para reportar altas/bajas/cambios."""
import openpyxl


def _col_map(path, col_a, col_b):
    """{valor_colA: valor_colB} por hoja, ubicando columnas por nombre de header.
    Usa iter_rows (funciona en read_only aunque max_column sea None por imágenes)."""
    m = {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        idx = {name: i for i, name in enumerate(header) if name is not None}
        ai = idx.get(col_a)
        if ai is None and col_a == "SKU":
            ai = idx.get("Número de artículo")
        bi = idx.get(col_b)
        if ai is None or bi is None:
            continue
        for row in it:
            a = row[ai] if ai < len(row) else None
            b = row[bi] if bi < len(row) else None
            if a is not None:
                m[str(a).strip()] = b
    wb.close()
    return m


def read_pedido_map(master_path):
    """{SKU: Pedido} del maestro, sin los vacíos (para no perder lo manual)."""
    return {k: v for k, v in _col_map(master_path, "SKU", "Pedido").items()
            if v not in (None, "")}


def stock_map(path):
    """{SKU: Disponibilidad} de un normalizado (para el diff de reportes)."""
    return _col_map(path, "SKU", "Disponibilidad")


def diff(old_master_path, new_master_path):
    """Reporta altas, bajas y cambios de cantidad entre dos maestros."""
    a = stock_map(old_master_path)
    b = stock_map(new_master_path)
    altas = [k for k in b if k not in a]
    bajas = [k for k in a if k not in b]
    cambios = [(k, a[k], b[k]) for k in b if k in a and a[k] != b[k]]
    return {"altas": altas, "bajas": bajas, "cambios": cambios}
