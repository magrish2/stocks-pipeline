#!/usr/bin/env python3
"""Actualiza los maestros de Drive al formato apto para Grisma, IN-PLACE
(preserva fotos, Pedido y columnas ocultas):
  - "Disponibilidad" -> "Disponible" y valores numéricos ('+240' -> 240).
  - Promos: "Precio mayorista" por par (módulo / pares) + columna "Descuento %".
Borra los maestros basura (vacíos).
"""
import os
import sys
import tempfile

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import drive
import json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import normalizar_stock as ns


def _headers(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def fix_master(path):
    """Devuelve (True, resumen) si lo modificó; (False, motivo) si es basura."""
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    if ws.max_row < 2 or ws.max_column < 5:
        return False, "vacío"

    h = _headers(ws)
    # 1) Disponibilidad -> Disponible + numérico
    disp_c = h.get("Disponibilidad") or h.get("Disponible")
    if h.get("Disponibilidad"):
        ws.cell(1, h["Disponibilidad"]).value = "Disponible"
    if disp_c:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, disp_c).value = ns.disp_value(ws.cell(r, disp_c).value)

    # 2) Promo: si hay columna oculta "Descuento" con valores
    desc_src = h.get("Descuento") or h.get("DESCUENTO")
    may_c = h.get("Precio mayorista")
    pub_c = h.get("Precio público")
    sku_c = h.get("SKU")
    n_promo = 0
    if desc_src and may_c and sku_c:
        # ¿algún descuento > 0? (si no, no es promo real)
        vals = [ns.descuento_pct(ws.cell(r, desc_src).value)
                for r in range(2, ws.max_row + 1)]
        if any(v for v in vals):
            # columna "Descuento %" nueva al final
            dc = ws.max_column + 1
            hdr = ws.cell(1, dc, "Descuento %")
            hdr.fill = PatternFill("solid", fgColor="1F2937")
            hdr.font = Font(color="FFFFFF", bold=True)
            hdr.alignment = Alignment(horizontal="center", vertical="center")
            for i, r in enumerate(range(2, ws.max_row + 1)):
                d = vals[i]
                if not d:
                    continue
                # por par = módulo con descuento / pares. Solo si el precio
                # parece un MÓDULO (mayor que el público) para no dividir precios
                # que ya vienen por par.
                pares = ns.get_pares(ws.cell(r, sku_c).value)
                cur = _num(ws.cell(r, may_c).value)
                pub = _num(ws.cell(r, pub_c).value) if pub_c else None
                if cur is not None and pares > 1 and (pub is None or cur > pub):
                    ws.cell(r, may_c).value = round(cur / pares, 2)
                cell = ws.cell(r, dc, d)
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="center")
                n_promo += 1

    wb.save(path)
    return True, f"disp={'ok' if disp_c else '-'} promo_rows={n_promo}"


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", ".").strip())
    except ValueError:
        return None


def main():
    cfg = json.load(open(os.path.join(os.path.dirname(__file__), "config.json")))
    svc = drive.service()
    tmp = tempfile.mkdtemp()
    remote = drive.list_files(svc, cfg["fijos"])
    for fid, name, mime in remote:
        if not name.lower().endswith(".xlsx") or name.startswith("~$"):
            continue
        d = os.path.join(tmp, name)
        drive.download(svc, fid, d, mime)
        try:
            ok, info = fix_master(d)
        except Exception as e:
            print(f"  [!] {name}: ERROR {e}")
            continue
        if not ok:
            drive.trash(svc, fid)
            print(f"  🗑️  BASURA borrado: {name} ({info})")
            continue
        drive.update_content(svc, fid, d)
        print(f"  ✅ {name}: {info}")
    print("\nListo.")


if __name__ == "__main__":
    main()
