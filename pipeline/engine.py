#!/usr/bin/env python3
"""Motor de normalización: envuelve normalizar_stock para Reebok y Kappa,
con soporte de sync de maestro (saltear sin stock + arrastrar Pedido)."""
import os
import re
import sys
from types import SimpleNamespace

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import normalizar_stock as ns
import normalizar_kappa as nk

REEBOK_CDN = "https://reebok.com.ar/cdn/shop/files/{name}?width=1024"
KAPPA_CDN = "https://www.kappastore.com.ar/cdn/shop/files/{name}?width=1024"


def _first_sku(path):
    """Primer SKU con dato del archivo (para detectar marca por prefijo)."""
    try:
        for sh in ns.list_sheets(path):
            rows = ns.read_sheet_rows(path, sh)
            h = ns.find_header_index(rows)
            if h is None:
                continue
            cmap = ns.col_map(rows[h])
            si = ns.pick(cmap, "sku", "número de artículo", "numero de articulo")
            if si is None:
                continue
            for r in rows[h + 1:]:
                if si < len(r) and r[si]:
                    return str(r[si]).strip().upper()
    except Exception:
        pass
    return ""


def detect_brand(path):
    n = os.path.basename(path).lower()
    if "kappa" in n:
        return "kappa"
    if "croc" in n:
        return "crocs"
    if "reebok" in n or "rbk" in n:
        return "reebok"
    # No está en el nombre (ej. stocks de clubes Kappa): mirar el prefijo del SKU.
    sku = _first_sku(path)
    if sku.startswith("RBK"):
        return "reebok"
    if re.match(r"^K\d", sku):
        return "kappa"
    if re.match(r"^C\d", sku):
        return "crocs"
    return "reebok"


def _sheets_for(path, brand):
    if brand == "kappa":
        return nk.stock_sheets(path)          # excluye BASE SKU / Hoja2 / (2)
    return None                                # Reebok: cmd_run elige CALZADO/INDUMENTARIA


def _apply_brand(brand, thumb_px=512):
    """Configura la fuente de imágenes (CDN/finder) según la marca."""
    ns.THUMB_PX = thumb_px
    ns.IMAGE_URL_FINDER = None
    if brand == "kappa":
        ns.REEBOK_CDN = KAPPA_CDN
    elif brand == "crocs":
        # Crocs no tiene un CDN "adivinable" al 100%: usamos finder multi-fuente.
        import requests
        import crocs
        sess = requests.Session()
        sess.headers.update({"User-Agent": "Mozilla/5.0"})
        finder, _n = crocs.make_finder(sess)
        ns.IMAGE_URL_FINDER = finder
    else:
        ns.REEBOK_CDN = REEBOK_CDN


def normalize(raw, out_path, skip_out_of_stock=False, carry=None, thumb_px=512):
    """Normaliza `raw` a `out_path`. Devuelve out_path."""
    brand = detect_brand(raw)
    _apply_brand(brand, thumb_px)
    args = SimpleNamespace(
        input=raw, output=out_path, source="reebok", no_images=False,
        image_mode="embed", online=True, limit=None,
        sheets=_sheets_for(raw, brand),
        skip_out_of_stock=skip_out_of_stock, carry=carry)
    ns.cmd_run(args)
    return out_path


# Columnas que delatan un archivo de "pendientes" (pedidos por cliente), no un stock.
PENDIENTE_MARKERS = {"a liberar", "nº documento", "n° documento",
                     "nro documento", "nombre sn"}


def is_pendiente(path):
    """True si el archivo es un listado de pendientes (no un stock)."""
    try:
        for sh in ns.list_sheets(path):
            rows = ns.read_sheet_rows(path, sh)
            for r in rows[:8]:
                names = {re.sub(r"\s+", " ", str(c).strip()).lower()
                         for c in r if c is not None}
                if names & PENDIENTE_MARKERS:
                    return True
    except Exception:
        pass
    return False


_SHOPIFY = {
    "reebok": "https://reebok.com.ar/cdn/shop/files/{name}?width=1024",
    "kappa": "https://www.kappastore.com.ar/cdn/shop/files/{name}?width=1024",
}
_SUFFIXES = ["-1.jpg", "_1.jpg", "-1.png"]


def _shopify_probe(base, model, sess):
    for suf in _SUFFIXES:
        url = base.format(name=str(model).strip() + suf)
        try:
            r = sess.head(url, timeout=10, allow_redirects=True)
            if r.status_code == 200 and \
                    r.headers.get("content-type", "").startswith("image"):
                return url
        except Exception:
            pass
    return None


def make_multibrand_finder():
    """finder(model_code, session) que resuelve la foto según el prefijo del SKU:
    RBK->reebok.com.ar, K#->kappastore, C#->crocs (multi-fuente)."""
    _crocs = {}

    def finder(model_code, session):
        up = str(model_code).strip().upper()
        if up.startswith("RBK"):
            return _shopify_probe(_SHOPIFY["reebok"], model_code, session)
        if re.match(r"^K\d", up):
            return _shopify_probe(_SHOPIFY["kappa"], model_code, session)
        if re.match(r"^C\d", up):
            if "f" not in _crocs:
                import crocs
                _crocs["f"], _ = crocs.make_finder(session)
            return _crocs["f"](model_code, session)
        return None

    return finder


def _pendiente_sku_col(path):
    """Nombre de la columna de SKU en un pendiente (header en la fila 1)."""
    import openpyxl
    ws = openpyxl.load_workbook(path, read_only=True).worksheets[0]
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(1, c).value
        if v and re.sub(r"\s+", " ", str(v).strip()).lower() in (
                "número", "numero", "sku", "número de artículo",
                "numero de articulo", "código", "codigo"):
            return str(v)
    return "Número"


def add_images_pendiente(raw, out_path, thumb_px=512):
    """Agrega columna Imagen a un pendiente (foto por SKU, marca por marca:
    puede mezclar Reebok/Kappa/Crocs). Conserva todo; no genera maestro."""
    ns.THUMB_PX = thumb_px
    ns.IMAGE_URL_FINDER = make_multibrand_finder()
    args = SimpleNamespace(input=raw, output=out_path,
                           sku_col=_pendiente_sku_col(raw),
                           source="reebok", fallback_stock=None, sheets=None)
    ns.cmd_add_images(args)
    ns.IMAGE_URL_FINDER = None
    return out_path
