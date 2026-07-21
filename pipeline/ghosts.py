#!/usr/bin/env python3
"""Arrastra a 0 los artículos que desaparecieron del crudo nuevo.

Compara el maestro viejo con el nuevo (recién regenerado desde el crudo) y, para
cada SKU que estaba en el viejo pero no en el nuevo, agrega una fila al final con
Disponibilidad = 0, conservando sus datos (descripción, precios, Pedido) y su
foto (re-descargada por Modelo-Color con la fuente ya configurada por el engine).
"""
import os
import sys

import openpyxl
import requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import normalizar_stock as ns
from openpyxl.styles import Alignment


def _headers(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def carry_zero(new_path, old_path, log=print):
    new_wb = openpyxl.load_workbook(new_path)                    # editable (con fotos)
    old_wb = openpyxl.load_workbook(old_path, data_only=True)    # sin read_only
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})
    cache = ns.load_json(ns.CACHE_PATH, default={})
    os.makedirs(ns.IMG_CACHE_DIR, exist_ok=True)
    row_h = round(ns.ROW_H_PX * 3 / 4, 1)
    old_by_title = {ws.title: ws for ws in old_wb.worksheets}
    total = 0

    for ws in new_wb.worksheets:
        old_ws = old_by_title.get(ws.title)
        if old_ws is None:
            continue
        hn = _headers(ws)
        ho = _headers(old_ws)
        sku_n = hn.get("SKU")
        sku_o = ho.get("SKU")
        if not sku_n or not sku_o:
            continue

        new_skus = {str(ws.cell(r, sku_n).value).strip()
                    for r in range(2, ws.max_row + 1)
                    if ws.cell(r, sku_n).value is not None}

        # columnas a copiar (las que existen en ambos), y la de foto (Modelo-Color)
        common = [h for h in hn if h in ho and h not in (None, "Foto")]
        mc_o = ho.get("Modelo-Color") or ho.get("Modelo Color") or ho.get("Modelo")

        at = ws.max_row + 1
        for r in range(2, old_ws.max_row + 1):
            sku = old_ws.cell(r, sku_o).value
            if sku is None or str(sku).strip() in new_skus:
                continue
            for h in common:
                val = old_ws.cell(r, ho[h]).value
                if h == "Disponible":
                    val = 0                      # el que desapareció -> 0
                cell = ws.cell(at, hn[h], value=val)
                cell.alignment = Alignment(vertical="center", wrap_text=(h in
                                           ("Descripción", "SKU")))
            # foto por Modelo-Color (fuente ya seteada por engine.normalize)
            mc = old_ws.cell(r, mc_o).value if mc_o else None
            if mc:
                key = str(mc).strip()
                p = ns.resolve_model_image(key, key, key, cache, session,
                                           source="reebok")
                if p:
                    ns.embed_centered(ws, p, at)
            ws.row_dimensions[at].height = row_h
            at += 1
            total += 1

    ns.save_json(ns.CACHE_PATH, cache)
    new_wb.save(new_path)
    log(f"  arrastrados en 0 (desaparecidos del crudo): {total}")
    return total
