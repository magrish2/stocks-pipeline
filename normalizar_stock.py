#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Normalizador de stock de calzado/indumentaria Reebok.

Lee un archivo de stock (.xlsb o .xlsx), busca la foto de cada modelo en
MercadoLibre Argentina usando la API oficial, y genera un Excel normalizado
con: Foto, SKU, Descripción, Disponibilidad, Precio mayorista, Precio público
y una columna de Pedido para completar.

Uso típico
----------
  1) Configurar credenciales (una sola vez):
        cp meli_config.example.json meli_config.json
        # editar meli_config.json con tu client_id / client_secret / redirect_uri

  2) Autorizar la app (una sola vez; el token se refresca solo después):
        python normalizar_stock.py auth-url           # abrí la URL, autorizá
        python normalizar_stock.py auth-exchange CODE  # pegá el code de la URL

  3) Generar el Excel normalizado:
        python normalizar_stock.py run
        # o con un archivo puntual:
        python normalizar_stock.py run --input "Stock REEBOK inmediato 18-06 +.xlsb"

Otros flags útiles de 'run':
    --no-images        No consultar MELI; deja la columna Foto vacía (rápido).
    --limit N          Procesar solo N filas por hoja (para pruebas).
    --sheets CALZADO   Procesar solo ciertas hojas (repetible).
    --image-mode       embed (default) | first-row | url
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
import time
import urllib.parse
from datetime import datetime
from io import BytesIO

import requests

# Dependencias de Excel / imagen
try:
    from pyxlsb import open_workbook as open_xlsb
except ImportError:
    open_xlsb = None

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None


# --------------------------------------------------------------------------- #
# Rutas y configuración
# --------------------------------------------------------------------------- #
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "meli_config.json")
TOKENS_PATH = os.path.join(BASE_DIR, "meli_tokens.json")
CACHE_PATH = os.path.join(BASE_DIR, "meli_cache.json")        # modelo -> url|null
IMG_CACHE_DIR = os.path.join(BASE_DIR, ".cache_img")
# Correcciones manuales modelo -> URL de imagen. Tienen prioridad sobre la URL
# adivinada del CDN (para modelos donde el nombre "limpio" trae otra foto).
IMG_OVERRIDES_PATH = os.path.join(BASE_DIR, "image_overrides.json")

# Endpoints de MercadoLibre Argentina
# Nota: el endpoint de items /sites/MLA/search está restringido (403 incluso con
# token válido). Usamos la API de catálogo de productos, que sí responde.
MELI_AUTH_URL = "https://auth.mercadolibre.com.ar/authorization"
MELI_TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
MELI_PRODSEARCH_URL = "https://api.mercadolibre.com/products/search"
MELI_PRODUCT_URL = "https://api.mercadolibre.com/products/{pid}"

# Cuántos productos del resultado revisar buscando uno que tenga foto.
MAX_PRODUCT_LOOKUPS = 8

# Tienda oficial Reebok Argentina (Shopify). La imagen del modelo es predecible:
#   https://reebok.com.ar/cdn/shop/files/{MODELO}-1.jpg?width=690
# donde {MODELO} es el código sin el talle (col "Modelo", ej. RBK1100227787).
# Es la mejor fuente: color exacto y alta resolución.
REEBOK_CDN = "https://reebok.com.ar/cdn/shop/files/{name}?width=690"
REEBOK_IMG_SUFFIXES = ["-1.jpg", "_1.jpg", "-1.png"]

# Encabezados esperados en la fila de títulos del stock
HEADER_ROW_HINT = "Número de artículo"
# Marcadores válidos para ubicar la fila de encabezados (Reebok usa "Número de
# artículo"; muchas planillas Kappa usan "SKU").
HEADER_HINTS = ("Número de artículo", "Numero de articulo", "SKU")

# Tamaño del thumbnail embebido (px). Más grande = mejor calidad, Excel más pesado.
THUMB_PX = 300
# Calidad JPEG del thumbnail embebido (1-95).
THUMB_QUALITY = 88


# --------------------------------------------------------------------------- #
# Utilidades de config / tokens / cache
# --------------------------------------------------------------------------- #
def load_json(path, default=None):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


def load_config():
    cfg = load_json(CONFIG_PATH)
    if not cfg:
        sys.exit(
            "ERROR: falta meli_config.json.\n"
            "  Copiá meli_config.example.json a meli_config.json y completá\n"
            "  client_id, client_secret y redirect_uri (los obtenés creando una\n"
            "  app gratis en https://developers.mercadolibre.com.ar/)."
        )
    for k in ("client_id", "client_secret", "redirect_uri"):
        if not cfg.get(k):
            sys.exit(f"ERROR: falta '{k}' en meli_config.json")
    return cfg


# --------------------------------------------------------------------------- #
# OAuth 2.0 de MercadoLibre
# --------------------------------------------------------------------------- #
def cmd_auth_url(_args):
    cfg = load_config()
    params = {
        "response_type": "code",
        "client_id": cfg["client_id"],
        "redirect_uri": cfg["redirect_uri"],
    }
    url = MELI_AUTH_URL + "?" + urllib.parse.urlencode(params)
    print("\n1) Abrí esta URL en el navegador (logueado con tu cuenta MELI):\n")
    print("   " + url)
    print(
        "\n2) Autorizá. El navegador te redirige a tu redirect_uri con\n"
        "   ?code=XXXXXXXX al final. Copiá ese code y ejecutá:\n\n"
        "   python normalizar_stock.py auth-exchange XXXXXXXX\n"
    )


def _save_tokens(tok):
    tok["obtained_at"] = int(time.time())
    save_json(TOKENS_PATH, tok)


def cmd_auth_exchange(args):
    cfg = load_config()
    data = {
        "grant_type": "authorization_code",
        "client_id": cfg["client_id"],
        "client_secret": cfg["client_secret"],
        "code": args.code,
        "redirect_uri": cfg["redirect_uri"],
    }
    r = requests.post(MELI_TOKEN_URL, data=data, timeout=30,
                      headers={"Accept": "application/json"})
    if r.status_code != 200:
        sys.exit(f"ERROR al canjear el code: {r.status_code} {r.text}")
    tok = r.json()
    _save_tokens(tok)
    print("OK: tokens guardados en meli_tokens.json. Ya podés correr 'run'.")


def _refresh_token(cfg, tokens):
    data = {
        "grant_type": "refresh_token",
        "client_id": cfg["client_id"],
        "client_secret": cfg["client_secret"],
        "refresh_token": tokens["refresh_token"],
    }
    r = requests.post(MELI_TOKEN_URL, data=data, timeout=30,
                      headers={"Accept": "application/json"})
    if r.status_code != 200:
        sys.exit(
            f"ERROR al refrescar token: {r.status_code} {r.text}\n"
            "Volvé a autorizar con auth-url / auth-exchange."
        )
    tok = r.json()
    _save_tokens(tok)
    return tok


def get_access_token(cfg):
    tokens = load_json(TOKENS_PATH)
    if not tokens:
        sys.exit(
            "ERROR: no hay tokens. Autorizá primero:\n"
            "  python normalizar_stock.py auth-url\n"
            "  python normalizar_stock.py auth-exchange CODE"
        )
    # expires_in suele ser 21600s (6h); refrescamos con margen de 5 min.
    age = int(time.time()) - tokens.get("obtained_at", 0)
    if age >= tokens.get("expires_in", 21600) - 300:
        tokens = _refresh_token(cfg, tokens)
    return tokens["access_token"]


# --------------------------------------------------------------------------- #
# Lectura del stock (.xlsb o .xlsx)
# --------------------------------------------------------------------------- #
def find_default_input():
    """Devuelve el archivo de stock más reciente que coincida con el patrón."""
    patterns = [
        os.path.join(BASE_DIR, "Stock REEBOK inmediato*.xls*"),
        os.path.join(os.path.expanduser("~/Desktop"), "Stock REEBOK inmediato*.xls*"),
        os.path.join(os.path.expanduser("~/Documents"), "Stock REEBOK inmediato*.xls*"),
    ]
    found = []
    for p in patterns:
        found.extend(glob.glob(p))
    found = [f for f in found if not os.path.basename(f).startswith("~$")]
    if not found:
        return None
    return max(found, key=os.path.getmtime)


def _rows_from_xlsb(path, sheet):
    with open_xlsb(path) as wb:
        if sheet not in wb.sheets:
            return None
        with wb.get_sheet(sheet) as ws:
            return [[c.v for c in row] for row in ws.rows()]


def _rows_from_xlsx(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def list_sheets(path):
    if path.lower().endswith(".xlsb"):
        if open_xlsb is None:
            sys.exit("ERROR: falta pyxlsb. Instalá con: pip install pyxlsb")
        with open_xlsb(path) as wb:
            return list(wb.sheets)
    wb = openpyxl.load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def read_sheet_rows(path, sheet):
    if path.lower().endswith(".xlsb"):
        return _rows_from_xlsb(path, sheet)
    return _rows_from_xlsx(path, sheet)


def extract_original_images(path):
    """
    Extrae las imágenes embebidas del archivo original (xlsb/xlsx, son ZIP).
    Devuelve {nombre_hoja: {fila_0based: (ext, bytes)}} mapeando por la columna
    de la imagen (col 0). Sirve de fallback cuando no hay foto online.
    """
    import zipfile
    from xml.etree import ElementTree as ET

    NS = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    R_EMBED = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"

    try:
        sheet_names = list_sheets(path)
    except Exception:
        return {}

    out = {}
    try:
        z = zipfile.ZipFile(path)
    except Exception:
        return {}
    names = set(z.namelist())

    with z:
        for i, sheet in enumerate(sheet_names, start=1):
            # sheet{i}.bin/.xml -> su rels -> drawing -> rels -> media
            base = f"xl/worksheets/sheet{i}"
            rels_name = f"{base}.bin.rels" if f"{base}.bin" in names else f"{base}.xml.rels"
            rels_name = f"xl/worksheets/_rels/{os.path.basename(rels_name)}"
            if rels_name not in names:
                continue
            try:
                srels = ET.fromstring(z.read(rels_name))
            except Exception:
                continue
            drawing_target = None
            for rel in srels:
                if rel.get("Type", "").endswith("/drawing"):
                    drawing_target = rel.get("Target")
            if not drawing_target:
                continue
            drawing_path = os.path.normpath(
                os.path.join("xl/worksheets", drawing_target)).replace("\\", "/")
            if drawing_path not in names:
                continue

            # rels del drawing: rId -> media
            drawing_rels = (f"xl/drawings/_rels/"
                            f"{os.path.basename(drawing_path)}.rels")
            rid2media = {}
            if drawing_rels in names:
                try:
                    drels = ET.fromstring(z.read(drawing_rels))
                    for rel in drels:
                        tgt = os.path.normpath(
                            os.path.join("xl/drawings", rel.get("Target"))
                        ).replace("\\", "/")
                        rid2media[rel.get("Id")] = tgt
                except Exception:
                    pass

            try:
                d = ET.fromstring(z.read(drawing_path))
            except Exception:
                continue
            anchors = (d.findall("xdr:twoCellAnchor", NS)
                       + d.findall("xdr:oneCellAnchor", NS))
            row_map = {}
            for a in anchors:
                frm = a.find("xdr:from", NS)
                if frm is None:
                    continue
                row_el = frm.find("xdr:row", NS)
                if row_el is None:
                    continue
                blip = a.find(".//a:blip", NS)
                if blip is None:
                    continue
                media = rid2media.get(blip.get(R_EMBED))
                if not media or media not in names:
                    continue
                try:
                    row0 = int(row_el.text)
                except (TypeError, ValueError):
                    continue
                ext = os.path.splitext(media)[1].lstrip(".").lower() or "png"
                row_map[row0] = (ext, z.read(media))
            if row_map:
                out[sheet] = row_map
    return out


def cell_images(path):
    """
    Extrae imágenes 'en celda' (rich values, feature de Excel "Insertar imagen
    en celda") de un xlsx. Devuelve {hoja: {fila_1based: (ext, bytes)}}.
    Cadena: celda vm -> metadata.xml -> richvalue -> richValueRel -> media.
    """
    import zipfile
    from xml.etree import ElementTree as ET

    RNS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    XLRD = "{http://schemas.microsoft.com/office/spreadsheetml/2017/richdata}"
    SNS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

    try:
        z = zipfile.ZipFile(path)
    except Exception:
        return {}
    names = set(z.namelist())
    if "xl/richData/richValueRel.xml" not in names:
        return {}
    with z:
        rels = ET.fromstring(z.read("xl/richData/_rels/richValueRel.xml.rels"))
        rid2media = {r.get("Id"): os.path.normpath(
            "xl/richData/" + r.get("Target")).replace("\\", "/") for r in rels}
        rvr = ET.fromstring(z.read("xl/richData/richValueRel.xml"))
        relidx2rid = [el.get(RNS + "id") for el in rvr]
        rd = ET.fromstring(z.read("xl/richData/rdrichvalue.xml"))
        rv2rel = []
        for rv in rd:
            vs = [v.text for v in rv]
            rv2rel.append(int(vs[0]) if vs and vs[0] is not None else None)
        md = ET.fromstring(z.read("xl/metadata.xml"))
        fut = []
        for fm in md.iter():
            if fm.tag.endswith("}futureMetadata"):
                for bk in fm:
                    rvb = bk.find(".//" + XLRD + "rvb")
                    fut.append(int(rvb.get("i")) if rvb is not None else None)
                break
        vm2J = []
        for vmd in md.iter():
            if vmd.tag.endswith("}valueMetadata"):
                for bk in vmd:
                    rc = [c for c in bk.iter() if c.tag.endswith("}rc")]
                    vm2J.append(int(rc[0].get("v")) if rc else None)
                break

        def media_for_vm(vm):
            try:
                rid = relidx2rid[rv2rel[fut[vm2J[vm - 1]]]]
                return rid2media.get(rid)
            except (IndexError, TypeError):
                return None

        wb = ET.fromstring(z.read("xl/workbook.xml"))
        sheets = [(s.get("name"), s.get(RNS + "id"))
                  for s in wb.iter() if s.tag.endswith("}sheet")]
        wrels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid2tgt = {r.get("Id"): r.get("Target") for r in wrels}

        out = {}
        for name, rid in sheets:
            tgt = rid2tgt.get(rid)
            if not tgt:
                continue
            wpath = os.path.normpath("xl/" + tgt).replace("\\", "/")
            if wpath not in names:
                continue
            sx = ET.fromstring(z.read(wpath))
            rowmap = {}
            for row in sx.iter(SNS + "row"):
                rnum = int(row.get("r"))
                for c in row:
                    vm = c.get("vm")
                    if vm:
                        m = media_for_vm(int(vm))
                        if m and m in names:
                            ext = os.path.splitext(m)[1].lstrip(".").lower()
                            rowmap[rnum] = (ext, z.read(m))
                            break
            if rowmap:
                out[name] = rowmap
    return out


def find_header_index(rows):
    """Ubica la fila de encabezados. La col de artículo se llama 'Número de
    artículo' (Reebok) o 'SKU' (varias planillas Kappa)."""
    for i, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(h.lower() in cells for h in HEADER_HINTS):
            return i
    return None


def col_map(header):
    """Mapea nombre de columna -> índice (case/espacios tolerante)."""
    m = {}
    for idx, name in enumerate(header):
        if name is None:
            continue
        key = re.sub(r"\s+", " ", str(name).strip()).lower()
        if key and key not in m:
            m[key] = idx
    return m


def pick(cmap, *candidates):
    for c in candidates:
        key = re.sub(r"\s+", " ", c.strip()).lower()
        if key in cmap:
            return cmap[key]
    return None


MODEL_CODE_RE = re.compile(r"^\s*([A-Z]{2}\d{3,5}|[A-Z0-9]{6})\s*-\s")

# Mapa de colores inglés (como vienen en la descripción Reebok) -> español
# (como aparece el atributo COLOR en el catálogo de MELI). Orden = prioridad.
COLOR_MAP = [
    ("white", "blanco"), ("ftwwht", "blanco"), ("chalk", "blanco"),
    ("black", "negro"), ("cblack", "negro"), ("core black", "negro"),
    ("grey", "gris"), ("gray", "gris"), ("pugry", "gris"),
    ("navy", "azul"), ("blue", "azul"), ("vecnav", "azul"),
    ("red", "rojo"), ("vectorred", "rojo"),
    ("green", "verde"),
    ("yellow", "amarillo"),
    ("pink", "rosa"), ("rose", "rosa"),
    ("purple", "violeta"), ("violet", "violeta"),
    ("orange", "naranja"),
    ("beige", "beige"), ("sand", "beige"), ("tan", "beige"),
    ("brown", "marron"),
    ("silver", "plateado"),
    ("gold", "dorado"),
]


def sku_color_es(descripcion, nombre_sin_talle):
    """Deduce el color principal del SKU (en español) a partir de la descripción."""
    if not descripcion:
        return None
    segs = [s.strip() for s in str(descripcion).split(" - ") if s.strip()]
    # El color suele ser el segmento anterior al talle (último). Probamos los
    # últimos segmentos por si el talle ocupa más de uno.
    candidates = " ".join(segs[-3:]).lower() if segs else ""
    # No mezclar con el nombre del modelo
    if nombre_sin_talle:
        candidates = candidates.replace(str(nombre_sin_talle).lower(), " ")
    for eng, esp in COLOR_MAP:
        if eng in candidates:
            return esp
    return None


def manufacturer_code(descripcion):
    """Extrae el código de fabricante (CN4107, GY0952, HS7769) si está al inicio."""
    if not descripcion:
        return None
    m = MODEL_CODE_RE.match(str(descripcion))
    if not m:
        return None
    code = m.group(1)
    # Evita falsos positivos tipo 'REEBOK' (6 letras sin dígitos)
    if not any(ch.isdigit() for ch in code):
        return None
    return code


def build_query(code, nombre_sin_talle, descripcion):
    """
    Arma el término de búsqueda para el catálogo de MELI. El nombre del modelo
    funciona; el código de fabricante (CN4107) NO está indexado en el catálogo,
    así que se prioriza el nombre.
    """
    if nombre_sin_talle:
        nombre = str(nombre_sin_talle).strip()
        if not nombre.lower().startswith("reebok"):
            nombre = "reebok " + nombre
        return nombre
    # Sin nombre: usar el primer segmento de la descripción que no sea el código
    if descripcion:
        segs = [s.strip() for s in str(descripcion).split(" - ") if s.strip()]
        for seg in segs:
            if seg != code and not MODEL_CODE_RE.match(seg + " - "):
                base = seg if seg.lower().startswith("reebok") else "reebok " + seg
                return base
    return "reebok"


# --------------------------------------------------------------------------- #
# Búsqueda de imágenes en MELI
# --------------------------------------------------------------------------- #
def _result_color(res):
    for a in (res.get("attributes") or []):
        if a.get("id") == "COLOR":
            return (a.get("value_name") or "").lower()
    return ""


def _pics_for(res, auth, session):
    """Devuelve las pictures de un resultado, pidiendo el detalle si hace falta."""
    pics = res.get("pictures") or []
    if pics:
        return pics
    pid = res.get("id")
    if not pid:
        return []
    try:
        d = session.get(MELI_PRODUCT_URL.format(pid=pid), headers=auth, timeout=20)
    except requests.RequestException:
        return []
    if d.status_code != 200:
        return []
    time.sleep(0.15)
    return d.json().get("pictures") or []


def meli_search_thumb(token, query, session, color_es=None):
    """
    Busca en el catálogo de MELI y devuelve la URL de una foto.
    Si se pasa color_es, prioriza el producto cuyo atributo COLOR coincide;
    si no encuentra ese color con foto, cae a la primera foto del modelo.
    """
    auth = {"Authorization": f"Bearer {token}"}
    try:
        r = session.get(
            MELI_PRODSEARCH_URL,
            params={"site_id": "MLA", "q": query, "limit": MAX_PRODUCT_LOOKUPS},
            headers=auth,
            timeout=20,
        )
    except requests.RequestException:
        return None
    if r.status_code != 200:
        return None
    results = r.json().get("results") or []
    if not results:
        return None

    # Ordenar: primero los del color buscado, manteniendo el orden original.
    if color_es:
        match = [x for x in results if color_es in _result_color(x)]
        rest = [x for x in results if color_es not in _result_color(x)]
        ordered = match + rest
    else:
        ordered = results

    for res in ordered:
        pics = _pics_for(res, auth, session)
        if pics:
            url = pics[0].get("url") or pics[0].get("secure_url")
            if url:
                return url.replace("http://", "https://")
    return None


def upgrade_url(url):
    """Lleva una URL de imagen de MELI a su versión original/grande (-O.jpg)."""
    if not url:
        return url
    url = url.replace("http://", "https://")
    # Quita prefijos de versión reducida (D_NQ_NP_, D_Q_NP_, D_NQ_NP_2X_) -> D_
    url = re.sub(r"/D_(?:[A-Z]+_NP_)?(?:\dX_)?", "/D_", url)
    # Cambia el sufijo de tamaño (-C/-F/-V/-I/-S/-N...) por -O (original)
    url = re.sub(r"-[A-Z]\.(jpg|jpeg|png|webp)$", r"-O.jpg", url, flags=re.I)
    return url


def download_thumb(url, dest_path, session):
    """Descarga y normaliza la imagen a un thumbnail. True si OK."""
    try:
        r = session.get(url, timeout=20)
    except requests.RequestException:
        return False
    if r.status_code != 200 or not r.content:
        return False
    if PILImage is None:
        with open(dest_path, "wb") as fh:
            fh.write(r.content)
        return True
    try:
        im = PILImage.open(BytesIO(r.content))
        # Si tiene transparencia (PNG/paleta), aplanar sobre BLANCO — si no, al
        # pasar a JPEG asoma el color "debajo" del alfa (p. ej. verde en Crocs).
        if im.mode in ("RGBA", "LA", "P"):
            im = im.convert("RGBA")
            bg = PILImage.new("RGB", im.size, (255, 255, 255))
            bg.paste(im, mask=im.split()[-1])
            im = bg
        else:
            im = im.convert("RGB")
        im.thumbnail((THUMB_PX, THUMB_PX))
        im.save(dest_path, "JPEG", quality=THUMB_QUALITY)
        return True
    except Exception:
        return False


def reebok_image_url(model_code, session):
    """
    Devuelve la URL de la foto oficial en reebok.com.ar para un modelo, o None.
    Prueba los nombres de archivo habituales del CDN de Shopify.
    """
    if not model_code:
        return None
    for suf in REEBOK_IMG_SUFFIXES:
        url = REEBOK_CDN.format(name=str(model_code).strip() + suf)
        try:
            r = session.head(url, timeout=12, allow_redirects=True)
        except requests.RequestException:
            continue
        if r.status_code == 200 and r.headers.get("content-type", "").startswith("image"):
            return url
    return None


_OVERRIDES = None

# Hook opcional: función finder(model_code, session) -> url|None que reemplaza
# al CDN adivinado (la usa Crocs, multi-fuente). None = comportamiento normal.
IMAGE_URL_FINDER = None


def image_overrides():
    """Mapa modelo_key -> URL corregida (image_overrides.json), cacheado."""
    global _OVERRIDES
    if _OVERRIDES is None:
        _OVERRIDES = load_json(IMG_OVERRIDES_PATH, default={}) or {}
    return _OVERRIDES


def resolve_model_image(modelo_key, model_code, query, cache, session,
                        token=None, color_es=None, source="both"):
    """
    Devuelve la ruta local de la foto de un modelo, usando cache en disco.
    Fuente: 'reebok' (web oficial), 'meli' (catálogo) o 'both' (Reebok y, si no
    está, MELI). cache: dict modelo_key -> url|None (persistido en meli_cache.json).
    """
    dest = os.path.join(IMG_CACHE_DIR, f"{re.sub(r'[^A-Za-z0-9_-]', '_', modelo_key)}.jpg")

    # Corrección manual: tiene prioridad sobre el disco y el CDN adivinado.
    ov = image_overrides().get(modelo_key)
    if ov:
        if download_thumb(ov, dest, session):
            return dest
        return dest if os.path.exists(dest) else None

    if os.path.exists(dest):
        return dest

    if modelo_key in cache:
        url = cache[modelo_key]
        if not url:
            return None
        if download_thumb(url, dest, session):
            return dest
        return None

    # No estaba en cache: resolver la URL según la fuente
    url = None
    if source in ("reebok", "both"):
        # Hook por marca (ej. Crocs multi-fuente); si no, CDN adivinado.
        if IMAGE_URL_FINDER is not None:
            url = IMAGE_URL_FINDER(model_code, session)
        else:
            url = reebok_image_url(model_code, session)
    if not url and source in ("meli", "both") and token:
        meli_url = meli_search_thumb(token, query, session, color_es=color_es)
        url = upgrade_url(meli_url) if meli_url else None
        time.sleep(0.2)  # cortesía de rate-limit MELI

    cache[modelo_key] = url
    if url and download_thumb(url, dest, session):
        return dest
    return None


# --------------------------------------------------------------------------- #
# Generación del Excel normalizado
# --------------------------------------------------------------------------- #
OUT_HEADERS = ["Foto", "SKU", "Descripción", "Género", "Disponibilidad",
               "Precio mayorista", "Mayorista + PP", "Precio público", "Pedido"]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILL = PatternFill("solid", fgColor="F7F8FA")        # filas pares (zebra)
PEDIDO_FILL = PatternFill("solid", fgColor="FFF7E6")       # resalta col Pedido
PRICE_FONT = Font(color="111827")
SKU_FONT = Font(bold=True, color="111827")
# Muestra el precio exacto del original (sin redondear). #,##0.## conserva los
# decimales solo si existen (59.999,5 / 119.999 / 29.804,85).
MONEY_FMT = '#,##0.##'

# Geometría de filas/columnas (en px para la foto; se traduce a unidades Excel).
ROW_H_PX = 116                 # alto de fila de datos
PHOTO_COL_PX = 132             # ancho de la columna Foto
PHOTO_BOX_PX = 104             # caja máxima de la imagen (deja margen)


def _emu_center_anchor(col0, row0, cell_w_px, cell_h_px, img_w_px, img_h_px):
    """
    Ancla una imagen centrada dentro de su celda usando twoCellAnchor en modo
    "mover con las celdas" (editAs=twoCell). Así la imagen se oculta/mueve junto
    con su fila al aplicar un filtro. Ambos marcadores quedan dentro de la misma
    celda (col0/row0), por lo que la imagen no se desborda.
    """
    c_off = max(0, (cell_w_px - img_w_px) / 2)
    r_off = max(0, (cell_h_px - img_h_px) / 2)
    m_from = AnchorMarker(col=col0, colOff=int(pixels_to_EMU(c_off)),
                          row=row0, rowOff=int(pixels_to_EMU(r_off)))
    m_to = AnchorMarker(col=col0, colOff=int(pixels_to_EMU(c_off + img_w_px)),
                        row=row0, rowOff=int(pixels_to_EMU(r_off + img_h_px)))
    return TwoCellAnchor(editAs="twoCell", _from=m_from, to=m_to)


def normalize_sheet(ws_out, rows, token, cache, session, opts, orig_images=None):
    hidx = find_header_index(rows)
    if hidx is None:
        print(f"  [!] No encontré la fila de encabezados; salteo la hoja.")
        return 0
    header = rows[hidx]
    cmap = col_map(header)

    i_sku = pick(cmap, "Número de artículo", "Numero de articulo", "SKU")
    i_desc = pick(cmap, "Descripción del artículo", "Descripcion del articulo",
                  "Descripción", "Descripcion",       # Crocs usa "Descripción" a secas
                  "ARTICLE NAME", "Nombre modelo",     # planillas de clubes Kappa
                  "Nombre del modelo", "Nombre del artículo")
    i_sintalle = pick(cmap, "Descripción sin talle", "Descripcion sin talle")
    i_modelo = pick(cmap, "Modelo Color", "Modelo-Color", "Modelo/Color", "Modelo")
    i_disp = pick(cmap, "DISPONIBLE (inmediato)", "Disponible (inmediato)")
    if i_disp is None:
        # El nombre varía por archivo: "DISPONIBLE", "DISPONIBLE (entrega 15
        # días)", "Disponible"... tomamos la primera columna que empiece así.
        for key, idx in cmap.items():
            # "DISPONIBLE …", "Stock (Inmediato)", "STOCK (entrega 15 días)"...
            if key.startswith("disponible") or key.startswith("stock"):
                i_disp = idx
                break
    # Prioridad: precio unitario "limpio"; si no hay (planillas de promo),
    # el "con descuento" (precio real de la promo); luego +PP y otros.
    i_may = pick(cmap, "Mayorista Unitario", "Mayorista con descuento",
                 "Mayorista con Descuento", "Mayorista unit. + PP",
                 "Mayorista", "Mayorista Primera Individual",
                 "Individual Mayorista")            # Crocs
    i_maypp = pick(cmap, "Mayorista unit. + PP", "Mayorista unit + PP",
                   "Mayorista Primera PP", "Módulo Mayorista")   # Crocs: módulo
    i_pub = pick(cmap, "Público", "Publico", "Individual Público",
                 "Individual Publico")              # Crocs
    i_genero = pick(cmap, "GÉNERO", "Género", "Genero", "Gender")

    if i_sku is None or i_desc is None:
        print("  [!] Faltan columnas mínimas (SKU/Descripción); salteo la hoja.")
        return 0

    # Encabezado de salida
    for c, name in enumerate(OUT_HEADERS, start=1):
        cell = ws_out.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws_out.row_dimensions[1].height = 30
    ws_out.freeze_panes = "B2"   # fija encabezado y la columna Foto

    # Anchos de columna (la Foto se calibra a PHOTO_COL_PX)
    photo_w = round((PHOTO_COL_PX - 5) / 7, 2)
    widths = {1: photo_w, 2: 24, 3: 50, 4: 10, 5: 15, 6: 16, 7: 16, 8: 16, 9: 11}
    for c, w in widths.items():
        ws_out.column_dimensions[get_column_letter(c)].width = w

    # Columnas originales que NO se muestran en el layout curado: en vez de
    # descartarlas, se anexan a la derecha OCULTAS (con su encabezado) para
    # poder mostrarlas si alguna vez se necesita esa info.
    i_img = pick(cmap, "Imagen", "FOTO", "Foto")
    consumed = {i_img, i_sku, i_desc, i_genero, i_disp, i_may, i_maypp, i_pub}
    extra_cols = []  # (idx_origen_0based, nombre)
    for j, name in enumerate(header):
        if name is None or j in consumed:
            continue
        if re.sub(r"\s+", " ", str(name).strip()).lower() == "pedido":
            continue  # ya hay una col Pedido en el layout
        extra_cols.append((j, str(name).strip()))
    extra_start = len(OUT_HEADERS) + 1
    for k, (j, name) in enumerate(extra_cols):
        c = extra_start + k
        cell = ws_out.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        col = ws_out.column_dimensions[get_column_letter(c)]
        col.width = 16
        col.hidden = True   # oculta; el usuario puede mostrarla cuando quiera

    out_row = 2
    seen_models = set()
    n_imgs = 0

    # Conservamos el índice original (0-based en la hoja) para mapear las
    # imágenes embebidas del archivo original (fallback).
    data_rows = [(hidx + 1 + j, r) for j, r in enumerate(rows[hidx + 1:])
                 if any(v is not None for v in r)]
    if opts.limit:
        data_rows = data_rows[: opts.limit]

    total = len(data_rows)
    for n, (orig_idx, row) in enumerate(data_rows, start=1):
        def g(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        sku = g(i_sku)
        desc = g(i_desc)
        if sku is None and desc is None:
            continue

        sintalle = g(i_sintalle)
        modelo = g(i_modelo) or sku
        genero = g(i_genero)
        disp = g(i_disp)
        may = g(i_may)
        maypp = g(i_maypp)
        pub = g(i_pub)

        # Sync de maestro: saltear lo que quedó sin stock.
        if getattr(opts, "skip_out_of_stock", False) and not in_stock(disp):
            continue

        c_sku = ws_out.cell(row=out_row, column=2, value=sku)
        ws_out.cell(row=out_row, column=3, value=desc)
        ws_out.cell(row=out_row, column=4, value=genero)
        ws_out.cell(row=out_row, column=5, value=disp)
        cmay = ws_out.cell(row=out_row, column=6, value=may)
        cmaypp = ws_out.cell(row=out_row, column=7, value=maypp)
        cpub = ws_out.cell(row=out_row, column=8, value=pub)
        c_sku.font = SKU_FONT
        for cell, val in ((cmay, may), (cmaypp, maypp), (cpub, pub)):
            cell.font = PRICE_FONT
            if isinstance(val, (int, float)):
                cell.number_format = MONEY_FMT
        # Pedido: a completar, salvo que se arrastre lo manual del maestro (sync).
        carry = getattr(opts, "carry", None)
        ped = carry.get(str(sku).strip()) if carry and sku is not None else None
        ws_out.cell(row=out_row, column=9, value=ped)

        band = BAND_FILL if (out_row % 2 == 0) else None
        for c in range(1, 10):
            cell = ws_out.cell(row=out_row, column=c)
            cell.border = BORDER
            if c == 9:
                cell.fill = PEDIDO_FILL        # columna Pedido siempre resaltada
            elif band is not None:
                cell.fill = band
            if c == 3:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif c == 2:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Columnas originales extra (ocultas): se copian tal cual del origen.
        for k, (j, _name) in enumerate(extra_cols):
            val = g(j)
            cell = ws_out.cell(row=out_row, column=extra_start + k, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(val, (int, float)):
                cell.number_format = MONEY_FMT

        # --- Foto ---
        if opts.image_mode != "off":
            code = manufacturer_code(desc)
            # Código de modelo para Reebok: la col "Modelo"; si no, el SKU sin talle.
            model_code = g(i_modelo)
            if not model_code and sku:
                model_code = re.sub(r"-[^-]+$", "", str(sku))
            model_key = str(modelo)
            query = build_query(code, sintalle, desc)
            first_of_model = model_key not in seen_models

            if opts.image_mode == "url":
                link = "https://reebok.com.ar/search?q=" + urllib.parse.quote(
                    str(model_code or query))
                cell = ws_out.cell(row=out_row, column=1, value="ver en Reebok")
                cell.hyperlink = link
                cell.font = Font(color="2563EB", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                want_img = (opts.image_mode == "embed") or first_of_model
                if want_img:
                    img_path = None
                    if getattr(opts, "online", True):
                        color_es = sku_color_es(desc, sintalle)
                        img_path = resolve_model_image(
                            model_key, model_code, query, cache, session,
                            token=token, color_es=color_es, source=opts.source)
                    # Fallback (y única fuente si online=False): la imagen que ya
                    # traía el archivo original en esa fila.
                    if not img_path and orig_images and orig_idx in orig_images:
                        ext, blob = orig_images[orig_idx]
                        op = os.path.join(
                            IMG_CACHE_DIR, f"orig_{ws_out.title}_{orig_idx}.{ext}")
                        if not os.path.exists(op):
                            with open(op, "wb") as fh:
                                fh.write(blob)
                        img_path = op
                    if img_path:
                        try:
                            xi = XLImage(img_path)
                            # Escala a la caja conservando proporción
                            w = h = PHOTO_BOX_PX
                            if PILImage is not None:
                                with PILImage.open(img_path) as _im:
                                    iw, ih = _im.size
                                scale = min(PHOTO_BOX_PX / iw, PHOTO_BOX_PX / ih)
                                w, h = int(iw * scale), int(ih * scale)
                            xi.width, xi.height = w, h
                            # Ancla centrada en la celda de la foto
                            xi.anchor = _emu_center_anchor(
                                0, out_row - 1, PHOTO_COL_PX, ROW_H_PX, w, h)
                            ws_out.add_image(xi)
                            n_imgs += 1
                        except Exception:
                            pass

        seen_models.add(str(modelo))
        ws_out.row_dimensions[out_row].height = round(ROW_H_PX * 3 / 4, 1)
        out_row += 1

        if n % 50 == 0 or n == total:
            print(f"    fila {n}/{total}  (fotos embebidas: {n_imgs})", flush=True)
            save_json(CACHE_PATH, cache)  # checkpoint del cache

    # Botones de filtro en el encabezado (sobre las columnas con datos).
    # Las fotos usan oneCellAnchor ("mover sin redimensionar"), así que se
    # ocultan junto con su fila cuando se aplica un filtro.
    last = out_row - 1
    if last >= 1:
        last_col = len(OUT_HEADERS) + len(extra_cols)
        ws_out.auto_filter.ref = f"B1:{get_column_letter(last_col)}{last}"

    return out_row - 2


def cmd_run(args):
    # Resolver archivo de entrada
    input_path = args.input or find_default_input()
    if not input_path or not os.path.exists(input_path):
        sys.exit("ERROR: no encontré archivo de stock. Pasá --input RUTA.")
    print(f"Archivo fuente: {input_path}")

    if input_path.lower().endswith(".xlsb") and open_xlsb is None:
        sys.exit("ERROR: falta pyxlsb para leer .xlsb. pip install pyxlsb")

    # Modo imagen
    image_mode = "off" if args.no_images else args.image_mode

    token = None
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
    })
    cache = load_json(CACHE_PATH, default={})

    if image_mode in ("embed", "first-row"):
        os.makedirs(IMG_CACHE_DIR, exist_ok=True)
        # Solo necesitamos credenciales de MELI si la fuente lo requiere.
        if args.source in ("meli", "both"):
            cfg = load_config()
            token = get_access_token(cfg)

    class Opts:
        pass
    opts = Opts()
    opts.limit = args.limit
    opts.image_mode = image_mode
    opts.source = args.source
    # online=False: no sondear la web (Kappa no tiene CDN); usar solo embebidas.
    opts.online = getattr(args, "online", True)
    # Sync de maestro: saltear sin stock y arrastrar Pedido manual por SKU.
    opts.skip_out_of_stock = getattr(args, "skip_out_of_stock", False)
    opts.carry = getattr(args, "carry", None)

    # Hojas a procesar
    all_sheets = list_sheets(input_path)
    target_sheets = args.sheets or [s for s in ("CALZADO", "INDUMENTARIA")
                                    if s in all_sheets]
    if not target_sheets:
        target_sheets = all_sheets
    print(f"Hojas a procesar: {target_sheets}")

    # Imágenes embebidas del original (fallback cuando no hay foto online).
    orig_all = {}
    if image_mode in ("embed", "first-row"):
        orig_all = extract_original_images(input_path)
        if orig_all:
            tot = sum(len(v) for v in orig_all.values())
            print(f"Imágenes en el archivo original (fallback): {tot}")

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    for sheet in target_sheets:
        print(f"\n== Hoja: {sheet} ==")
        rows = read_sheet_rows(input_path, sheet)
        if not rows:
            print("  (vacía o inexistente)")
            continue
        ws_out = out_wb.create_sheet(title=sheet[:31])
        n = normalize_sheet(ws_out, rows, token, cache, session, opts,
                            orig_images=orig_all.get(sheet))
        print(f"  -> {n} filas normalizadas")

    if not out_wb.sheetnames:
        sys.exit("No se generó ninguna hoja. Revisá el archivo de entrada.")

    save_json(CACHE_PATH, cache)
    fecha = datetime.now().strftime("%Y-%m-%d")
    out_name = args.output or os.path.join(
        BASE_DIR, f"Stock_REEBOK_normalizado_{fecha}.xlsx")
    out_wb.save(out_name)
    print(f"\n✅ Listo: {out_name}")


def in_stock(disp):
    """True si la disponibilidad indica que hay stock. Acepta números,
    '0'/vacío (sin stock) y strings tope tipo '+ 240' (hay stock)."""
    if disp is None:
        return False
    if isinstance(disp, (int, float)):
        return disp > 0
    s = str(disp).strip()
    if not s:
        return False
    if s.startswith("+"):
        return True
    num = re.sub(r"[^\d,.-]", "", s).replace(".", "").replace(",", ".")
    try:
        return float(num) > 0
    except ValueError:
        return True  # texto no vacío no numérico -> lo consideramos con stock


def model_from_sku(sku):
    """Código de modelo para Reebok a partir de un SKU (saca el talle y guiones)."""
    if not sku:
        return None
    return re.sub(r"-[^-]+$", "", str(sku).strip()).rstrip("-") or None


def embed_centered(ws, img_path, row1, col1=1,
                   col_px=PHOTO_COL_PX, row_px=ROW_H_PX, box_px=PHOTO_BOX_PX):
    """Embebe una imagen centrada en (row1, col1) (1-based). Devuelve True si OK."""
    try:
        xi = XLImage(img_path)
        w = h = box_px
        if PILImage is not None:
            with PILImage.open(img_path) as im:
                iw, ih = im.size
            scale = min(box_px / iw, box_px / ih)
            w, h = int(iw * scale), int(ih * scale)
        xi.width, xi.height = w, h
        xi.anchor = _emu_center_anchor(col1 - 1, row1 - 1, col_px, row_px, w, h)
        ws.add_image(xi)
        return True
    except Exception:
        return False


def build_columbia_index(catalog_paths):
    """
    Cruza las grillas Columbia (F26): imagen-en-celda -> clave 'style+color'
    (col C MATERIAL STYLE + col D Color Code). Devuelve {clave10dig: (ext,bytes)}.
    """
    index = {}
    for f in catalog_paths:
        try:
            ci = cell_images(f)
        except Exception:
            continue
        if not ci:
            continue
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            continue
        for sheet, rowmap in ci.items():
            if not sheet.upper().startswith("COLUMBIA"):
                continue
            ws = wb[sheet]
            for r, blob in rowmap.items():
                style = ws.cell(r, 3).value   # C: MATERIAL STYLE
                color = ws.cell(r, 4).value   # D: Color Code
                if style and color is not None:
                    key = str(style).strip() + str(color).strip().zfill(3)
                    index.setdefault(key, blob)
        wb.close()
    return index


def cmd_columbia_images(args):
    """Genera un xlsx del Database Columbia con columna Imagen, cruzando F26."""
    input_path = args.input
    if not input_path or not os.path.exists(input_path):
        sys.exit("ERROR: pasá --input con el .xlsb Columbia.")
    catalog_dir = args.catalogs or BASE_DIR  # raíz del proyecto (incluye F26)
    grillas = sorted(glob.glob(os.path.join(catalog_dir, "**", "*GRILLA*COLUMBIA*.xlsx"),
                               recursive=True))
    grillas = [g for g in grillas if not os.path.basename(g).startswith("~$")]
    if not grillas:
        sys.exit(f"ERROR: no encontré grillas COLUMBIA en {catalog_dir}")
    print(f"Grillas: {[os.path.basename(g) for g in grillas]}")
    print("Construyendo índice de imágenes (cruce por Material)...")
    index = build_columbia_index(grillas)
    print(f"  imágenes indexadas: {len(index)}")

    os.makedirs(IMG_CACHE_DIR, exist_ok=True)
    sheet = args.sheet or "Database"
    rows = read_sheet_rows(input_path, sheet)
    if not rows:
        sys.exit(f"ERROR: no pude leer la hoja '{sheet}'.")

    # Header: fila que contiene la columna del SKU (Material)
    sku_name = (args.sku_col or "Material").strip().lower()
    hidx = None
    for i, row in enumerate(rows[:10]):
        cells = [re.sub(r"\s+", " ", str(c).strip()).lower() if c is not None else ""
                 for c in row]
        if sku_name in cells:
            hidx = i
            break
    if hidx is None:
        sys.exit(f"ERROR: no encontré la columna '{args.sku_col}' en '{sheet}'.")
    header = [str(c) if c is not None else "" for c in rows[hidx]]
    mi = [j for j, c in enumerate(header) if c.strip().lower() == sku_name][0]
    date_cols = [j for j, c in enumerate(header) if "fecha" in c.lower()]

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = sheet[:31]
    out_headers = ["Imagen"] + header
    for c, name in enumerate(out_headers, start=1):
        cell = ws.cell(1, c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"
    photo_w = round((PHOTO_COL_PX - 5) / 7, 2)
    ws.column_dimensions["A"].width = photo_w
    row_h = round(ROW_H_PX * 3 / 4, 1)

    out_row = 2
    n_imgs = 0
    data = [r for r in rows[hidx + 1:] if any(v is not None for v in r)]
    for n, row in enumerate(data, start=1):
        for j, val in enumerate(header):
            v = row[j] if j < len(row) else None
            cell = ws.cell(out_row, j + 2, value=v)
            if j in date_cols and isinstance(v, (int, float)):
                cell.number_format = "dd/mm/yyyy"
        # clave Material (10 dígitos)
        mat = row[mi] if mi < len(row) else None
        key = None
        if mat is not None:
            try:
                key = str(int(float(mat)))
            except (TypeError, ValueError):
                key = str(mat).strip()
        if key and key in index:
            ext, blob = index[key]
            dest = os.path.join(IMG_CACHE_DIR, f"col_{key}.{ext}")
            if not os.path.exists(dest):
                with open(dest, "wb") as fh:
                    fh.write(blob)
            if embed_centered(ws, dest, out_row):
                n_imgs += 1
        ws.row_dimensions[out_row].height = row_h
        out_row += 1
        if n % 100 == 0 or n == len(data):
            print(f"    fila {n}/{len(data)} (fotos: {n_imgs})", flush=True)

    last = out_row - 1
    ws.auto_filter.ref = f"B1:{get_column_letter(len(out_headers))}{last}"

    out_name = args.output or os.path.join(
        os.path.dirname(os.path.abspath(input_path)),
        os.path.splitext(os.path.basename(input_path))[0] + " CON IMAGENES.xlsx")
    out_wb.save(out_name)
    print(f"\n✅ Listo ({n_imgs}/{len(data)} fotos): {out_name}")


def build_stock_image_index(paths):
    """
    Índice modelo -> (ext, bytes) a partir de las imágenes embebidas de planillas
    de stock Reebok (columna Imagen por fila). Clave = col 'Modelo' (o el SKU).
    Sirve para rellenar fotos que no están en reebok.com.ar.
    """
    index = {}
    for f in paths:
        if not os.path.exists(f):
            continue
        imgs = extract_original_images(f)
        for sheet, rowmap in imgs.items():
            rows = read_sheet_rows(f, sheet)
            if not rows:
                continue
            hidx = find_header_index(rows)
            if hidx is None:
                continue
            cmap = col_map(rows[hidx])
            mi = pick(cmap, "Modelo Color", "Modelo")
            si = pick(cmap, "Número de artículo", "Numero de articulo")
            for row0, blob in rowmap.items():
                if row0 >= len(rows):
                    continue
                r = rows[row0]
                mod = (r[mi] if mi is not None and mi < len(r) else None) \
                    or (r[si] if si is not None and si < len(r) else None)
                if mod:
                    index.setdefault(str(mod).strip(), blob)
    return index


def cmd_add_images(args):
    """Inserta una columna 'Imagen' al inicio de una planilla, sin tocar el resto."""
    input_path = args.input
    if not input_path or not os.path.exists(input_path):
        sys.exit("ERROR: pasá --input con la ruta de la planilla.")
    print(f"Planilla: {input_path}")

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
    })
    cache = load_json(CACHE_PATH, default={})
    os.makedirs(IMG_CACHE_DIR, exist_ok=True)

    token = None
    if args.source in ("meli", "both"):
        cfg = load_config()
        token = get_access_token(cfg)

    # Índice de fallback: imágenes embebidas de planillas de stock Reebok.
    stock_idx = {}
    if args.fallback_stock:
        stock_idx = build_stock_image_index(args.fallback_stock)
        print(f"Fallback de stock: {len(stock_idx)} modelos con imagen embebida")

    wb = openpyxl.load_workbook(input_path)  # conserva valores/estilos/fórmulas
    target = args.sheets or wb.sheetnames
    photo_w = round((PHOTO_COL_PX - 5) / 7, 2)
    row_h = round(ROW_H_PX * 3 / 4, 1)
    total_imgs = 0

    for sheet in target:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # Buscar la columna del SKU en la fila de encabezados (fila 1)
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        sku_col = None
        wanted = (args.sku_col or "Número").strip().lower()
        for c, name in enumerate(header, start=1):
            if name and re.sub(r"\s+", " ", str(name).strip()).lower() == wanted:
                sku_col = c
                break
        if sku_col is None:
            # fallback: cualquier columna que contenga 'numero'
            for c, name in enumerate(header, start=1):
                if name and "numero" in re.sub(r"[^a-z]", "", str(name).lower()):
                    sku_col = c
                    break
        if sku_col is None:
            print(f"  [!] {sheet}: no encontré la columna de SKU ('{args.sku_col}'); salteo.")
            continue

        # Insertar columna nueva al principio (todo se corre una posición)
        ws.insert_cols(1)
        sku_col += 1
        ws.cell(1, 1, value="Imagen")
        ws.cell(1, 1).font = Font(bold=True)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(1)].width = photo_w

        max_row = ws.max_row
        n_imgs = 0
        for r in range(2, max_row + 1):
            sku = ws.cell(r, sku_col).value
            model = model_from_sku(sku)
            if not model:
                continue
            img_path = resolve_model_image(
                model, model, model, cache, session,
                token=token, color_es=None, source=args.source)
            # Fallback: imagen embebida del stock para los que no están online
            if not img_path and model in stock_idx:
                ext, blob = stock_idx[model]
                sp = os.path.join(IMG_CACHE_DIR,
                                  f"stk_{re.sub(r'[^A-Za-z0-9_-]', '_', model)}.{ext}")
                if not os.path.exists(sp):
                    with open(sp, "wb") as fh:
                        fh.write(blob)
                img_path = sp
            if img_path and embed_centered(ws, img_path, r):
                n_imgs += 1
            ws.row_dimensions[r].height = row_h
            if (r - 1) % 50 == 0 or r == max_row:
                print(f"    {sheet}: fila {r-1}/{max_row-1} (fotos: {n_imgs})",
                      flush=True)
                save_json(CACHE_PATH, cache)
        total_imgs += n_imgs
        print(f"  -> {sheet}: {n_imgs} fotos agregadas")

    save_json(CACHE_PATH, cache)
    out_name = args.output or os.path.join(
        os.path.dirname(os.path.abspath(input_path)),
        os.path.splitext(os.path.basename(input_path))[0] + " CON IMAGENES.xlsx")
    wb.save(out_name)
    print(f"\n✅ Listo ({total_imgs} fotos): {out_name}")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def main():
    p = argparse.ArgumentParser(description="Normalizador de stock Reebok + fotos MELI")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("auth-url", help="Imprime la URL de autorización de MELI")

    pe = sub.add_parser("auth-exchange", help="Canjea el code por tokens")
    pe.add_argument("code", help="El valor 'code' de la URL de redirección")

    pr = sub.add_parser("run", help="Genera el Excel normalizado")
    pr.add_argument("--input", help="Ruta al .xlsb/.xlsx de stock")
    pr.add_argument("--output", help="Ruta del Excel de salida")
    pr.add_argument("--source", choices=["reebok", "meli", "both"],
                    default="reebok",
                    help="reebok: web oficial (color exacto, alta calidad) | "
                         "meli: catálogo MercadoLibre | both: Reebok y, si falta, MELI")
    pr.add_argument("--no-images", action="store_true",
                    help="No descargar fotos; columna Foto vacía")
    pr.add_argument("--no-online", dest="online", action="store_false",
                    help="No sondear la web; usar solo imágenes embebidas (Kappa)")
    pr.add_argument("--image-mode", choices=["embed", "first-row", "url"],
                    default="embed",
                    help="embed: foto en cada fila | first-row: solo 1ra fila por "
                         "modelo | url: hipervínculo de búsqueda")
    pr.add_argument("--limit", type=int, help="Procesar solo N filas por hoja")
    pr.add_argument("--sheets", action="append",
                    help="Hoja(s) a procesar (repetible)")

    pa = sub.add_parser("add-images",
                        help="Agrega una columna 'Imagen' al inicio de una planilla")
    pa.add_argument("--input", required=True, help="Ruta a la planilla .xlsx")
    pa.add_argument("--output", help="Ruta de salida (default: '<nombre> CON IMAGENES.xlsx')")
    pa.add_argument("--sku-col", default="Número",
                    help="Nombre de la columna que tiene el SKU (default: 'Número')")
    pa.add_argument("--source", choices=["reebok", "meli", "both"],
                    default="reebok", help="Fuente de las fotos (default: reebok)")
    pa.add_argument("--fallback-stock", action="append",
                    help="Planilla(s) de stock con imágenes embebidas para rellenar "
                         "lo que no esté en reebok.com.ar (repetible)")
    pa.add_argument("--sheets", action="append",
                    help="Hoja(s) a procesar (default: todas)")

    pc = sub.add_parser("columbia-images",
                        help="Genera el Database Columbia con imágenes cruzando F26")
    pc.add_argument("--input", required=True, help="Ruta al .xlsb Columbia")
    pc.add_argument("--catalogs", help="Carpeta con las grillas (default: ./F26)")
    pc.add_argument("--output", help="Ruta de salida")
    pc.add_argument("--sheet", default="Database", help="Hoja a procesar (default: Database)")
    pc.add_argument("--sku-col", default="Material",
                    help="Columna con el código (default: 'Material')")

    args = p.parse_args()
    if args.cmd == "auth-url":
        cmd_auth_url(args)
    elif args.cmd == "auth-exchange":
        cmd_auth_exchange(args)
    elif args.cmd == "run":
        cmd_run(args)
    elif args.cmd == "add-images":
        cmd_add_images(args)
    elif args.cmd == "columbia-images":
        cmd_columbia_images(args)


if __name__ == "__main__":
    main()
