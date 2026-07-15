#!/usr/bin/env python3
"""Winter Sale: toma la hoja 'Tabla Unica' TAL CUAL (todas sus columnas y
precios), saca filas vacías y le agrega una columna Imagen (Columbia) al inicio."""
import os
import re
import sys

import requests
sys.path.insert(0, "pipeline")
import normalizar_stock as ns
import engine
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = "WINTER SALE - Propuesta Mayorista.xlsx"
SHEET = "Tabla Unica"

ns.THUMB_PX = 512
ns.IMAGE_URL_FINDER = engine.make_columbia_finder()
cache = ns.load_json(ns.CACHE_PATH, default={})
os.makedirs(ns.IMG_CACHE_DIR, exist_ok=True)
sess = requests.Session()
sess.headers.update({"User-Agent": "Mozilla/5.0 (Macintosh) AppleWebKit/537.36 "
                     "Chrome/124 Safari/537.36"})

rows = ns.read_sheet_rows(SRC, SHEET)
hidx = ns.find_header_index(rows)
if hidx is None:
    for r in range(min(10, len(rows))):
        if sum(1 for c in rows[r] if c not in (None, "")) >= 5:
            hidx = r
            break
header = [c for c in rows[hidx]]
data = [r for r in rows[hidx + 1:] if any(v not in (None, "") for v in r)]
mc_idx = [i for i, c in enumerate(header)
          if str(c).strip().lower() == "modelo-color"][0]
print(f"columnas={len(header)} | filas con dato={len(data)}", flush=True)

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Winter Sale"
HF = PatternFill("solid", fgColor="1F4E5F"); HFONT = Font(color="FFFFFF", bold=True)
# Encabezado: Imagen + columnas originales
out_header = ["Imagen"] + [str(c).strip() if c is not None else "" for c in header]
for c, name in enumerate(out_header, 1):
    cell = ws.cell(1, c, name); cell.fill = HF; cell.font = HFONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30
ws.freeze_panes = "C2"
ws.column_dimensions["A"].width = round((ns.PHOTO_COL_PX - 5) / 7, 2)

n_img = 0
for i, r in enumerate(data, start=2):
    for c, v in enumerate(r, start=2):        # datos originales desde la col 2
        cell = ws.cell(i, c, v)
        if isinstance(v, (int, float)):
            cell.number_format = "#,##0.##"
        cell.alignment = Alignment(vertical="center")
    mc = r[mc_idx] if mc_idx < len(r) else None
    if mc:
        model = re.sub(r"-+$", "", str(mc).strip())
        p = ns.resolve_model_image(model, model, model, cache, sess, source="reebok")
        if p and ns.embed_centered(ws, p, i):
            n_img += 1
    ws.row_dimensions[i].height = round(ns.ROW_H_PX * 3 / 4, 1)
    if (i - 1) % 100 == 0:
        print(f"  {i-1}/{len(data)} (fotos {n_img})", flush=True)
        ns.save_json(ns.CACHE_PATH, cache)

last_col = get_column_letter(len(out_header))
ws.auto_filter.ref = f"B1:{last_col}{len(data)+1}"
ns.save_json(ns.CACHE_PATH, cache)
out = "WINTER SALE - Tabla con imagenes.xlsx"
wb.save(out)
print(f"LISTO: {len(data)} filas, {n_img} fotos -> {out}", flush=True)
