#!/usr/bin/env python3
"""Rellena la columna 'Imagen' existente de una planilla de stock (por Modelo),
sin insertar columnas ni tocar el resto. Maneja encabezado en cualquier fila.

Uso:
  python fill_images.py "<archivo>.xlsx" [--output OUT] [--fallback-stock X.xlsx ...]
"""
import argparse
import os
import re
import sys

import requests
import openpyxl
from openpyxl.utils import get_column_letter

import normalizar_stock as ns


def find_cols(ws):
    """Devuelve (header_row, imagen_col, modelo_col, sku_col) 1-based, o None."""
    for r in range(1, min(ws.max_row, 15) + 1):
        names = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                names[re.sub(r"\s+", " ", str(v).strip()).lower()] = c
        if ns.HEADER_ROW_HINT.lower() in names:
            img = names.get("imagen")
            mod = names.get("modelo color") or names.get("modelo")
            sku = names.get("número de artículo") or names.get("numero de articulo")
            return r, img, mod, sku
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--output")
    ap.add_argument("--fallback-stock", action="append", default=[])
    ap.add_argument("--source", default="reebok",
                    choices=["reebok", "meli", "both"])
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit(f"No existe: {args.input}")

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac "
                            "OS X 10_15_7) AppleWebKit/537.36 (KHTML, like "
                            "Gecko) Chrome/124 Safari/537.36"})
    cache = ns.load_json(ns.CACHE_PATH, default={})
    os.makedirs(ns.IMG_CACHE_DIR, exist_ok=True)

    token = None
    if args.source in ("meli", "both"):
        token = ns.get_access_token(ns.load_config())

    stock_idx = {}
    if args.fallback_stock:
        stock_idx = ns.build_stock_image_index(args.fallback_stock)
        print(f"Fallback de stock: {len(stock_idx)} modelos con imagen embebida")

    wb = openpyxl.load_workbook(args.input)
    photo_w = round((ns.PHOTO_COL_PX - 5) / 7, 2)
    row_h = round(ns.ROW_H_PX * 3 / 4, 1)
    total = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        found = find_cols(ws)
        if not found:
            print(f"  [!] {sheet}: sin encabezado; salteo.")
            continue
        hrow, img_c, mod_c, sku_c = found
        if img_c is None:
            print(f"  [!] {sheet}: sin columna 'Imagen'; salteo.")
            continue
        ws.column_dimensions[get_column_letter(img_c)].width = photo_w
        # Borrar imágenes embebidas previas (evita fotos viejas debajo de las nuevas)
        if getattr(ws, "_images", None):
            ws._images = []

        n = 0
        for r in range(hrow + 1, ws.max_row + 1):
            model = None
            if mod_c:
                model = ws.cell(r, mod_c).value
            if not model and sku_c:
                model = ns.model_from_sku(ws.cell(r, sku_c).value)
            model = str(model).strip() if model else None
            if not model:
                continue

            path = ns.resolve_model_image(model, model, model, cache, session,
                                          token=token, color_es=None,
                                          source=args.source)
            if not path and model in stock_idx:
                ext, blob = stock_idx[model]
                sp = os.path.join(ns.IMG_CACHE_DIR,
                                  f"stk_{re.sub(r'[^A-Za-z0-9_-]', '_', model)}.{ext}")
                if not os.path.exists(sp):
                    with open(sp, "wb") as fh:
                        fh.write(blob)
                path = sp

            if path and ns.embed_centered(ws, path, r, col1=img_c):
                n += 1
            ws.row_dimensions[r].height = row_h
            if (r - hrow) % 50 == 0 or r == ws.max_row:
                print(f"    {sheet}: {r-hrow}/{ws.max_row-hrow} (fotos: {n})",
                      flush=True)
                ns.save_json(ns.CACHE_PATH, cache)
        total += n
        print(f"  -> {sheet}: {n} fotos")

    ns.save_json(ns.CACHE_PATH, cache)
    out = args.output or (os.path.splitext(args.input)[0] + " CON IMAGENES.xlsx")
    wb.save(out)
    print(f"\n✅ Listo ({total} fotos): {out}")


if __name__ == "__main__":
    main()
