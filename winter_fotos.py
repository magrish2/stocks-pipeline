#!/usr/bin/env python3
"""Reemplaza SOLO las fotos de las hojas de presentación del Winter Sale por las
buenas de columbiasportswear, sin tocar nada más del workbook (pivots, fórmulas,
formato). Trabaja a nivel de bytes dentro del .xlsx."""
import io
import os
import sys
import zipfile
import xml.etree.ElementTree as ET

import requests
from PIL import Image
from openpyxl import load_workbook

sys.path.insert(0, "pipeline")
import engine

SRC = "WINTER SALE - Propuesta Mayorista.xlsx"
OUT = "WINTER SALE - Propuesta Mayorista - FOTOS.xlsx"
SHEETS = {"INDUMENTARIA": 7, "CALZADO": 8, "ACCESORIOS": 8}   # fila del encabezado
MODELO_COL = 2                                                # col B

R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A = "http://schemas.openxmlformats.org/drawingml/2006/main"
WNS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def norm(p):
    return os.path.normpath(p).replace("\\", "/")


# 1) row(0-based) -> Modelo por hoja
wb = load_workbook(SRC, read_only=True)
row_modelo = {}
for name, hr in SHEETS.items():
    ws = wb[name]; m = {}
    for ridx, vals in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True),
                                start=hr + 1):
        mod = vals[MODELO_COL - 1] if len(vals) >= MODELO_COL else None
        if mod and str(mod).startswith("COL"):
            m[ridx - 1] = str(mod).strip()      # row0 (0-based) del anchor
    row_modelo[name] = m
    print(f"{name}: {len(m)} filas con Modelo", flush=True)

z = zipfile.ZipFile(SRC)
names = set(z.namelist())
wbx = ET.fromstring(z.read("xl/workbook.xml"))
name_rid = {s.get("name"): s.get(f"{{{R}}}id")
            for s in wbx.find(f"{{{WNS}}}sheets")}
wbrels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
rid_target = {rel.get("Id"): rel.get("Target") for rel in wbrels}


def sheet_drawing(sheet_target):
    base = os.path.basename(sheet_target)
    relp = f"xl/worksheets/_rels/{base}.rels"
    if relp not in names:
        return None
    for rel in ET.fromstring(z.read(relp)):
        if rel.get("Type").endswith("/drawing"):
            return norm("xl/worksheets/" + rel.get("Target"))
    return None


finder = engine.make_columbia_finder()
sess = requests.Session()
sess.headers.update({"User-Agent": "Mozilla/5.0 (Macintosh) AppleWebKit/537.36 "
                     "Chrome/124 Safari/537.36"})
_img_cache = {}


def columbia_bytes(modelo, ext):
    key = (modelo, ext)
    if key in _img_cache:
        return _img_cache[key]
    out = None
    url = finder(modelo, sess)
    if url:
        try:
            r = sess.get(url, timeout=20)
            if r.status_code == 200 and r.content:
                im = Image.open(io.BytesIO(r.content))
                if im.mode in ("RGBA", "LA", "P"):
                    im = im.convert("RGBA")
                    bg = Image.new("RGB", im.size, (255, 255, 255))
                    bg.paste(im, mask=im.split()[-1]); im = bg
                else:
                    im = im.convert("RGB")
                im.thumbnail((700, 700))
                buf = io.BytesIO()
                im.save(buf, "PNG" if ext == "png" else "JPEG", quality=90)
                out = buf.getvalue()
        except Exception:
            out = None
    _img_cache[key] = out
    return out


# 2) media -> nuevos bytes
media_new = {}
n_ok = n_miss = 0
for name in SHEETS:
    target = rid_target[name_rid[name]]
    drawing = sheet_drawing(target)
    if not drawing or drawing not in names:
        continue
    dxml = ET.fromstring(z.read(drawing))
    drel = ET.fromstring(z.read(f"xl/drawings/_rels/{os.path.basename(drawing)}.rels"))
    rid_media = {rel.get("Id"): norm("xl/drawings/" + rel.get("Target"))
                 for rel in drel}
    for anchor in list(dxml):
        frm = anchor.find(f"{{{XDR}}}from")
        blip = anchor.find(f".//{{{A}}}blip")
        if frm is None or blip is None:
            continue
        row0 = int(frm.find(f"{{{XDR}}}row").text)
        media = rid_media.get(blip.get(f"{{{R}}}embed"))
        modelo = row_modelo[name].get(row0)
        if not media or not modelo:
            continue
        ext = os.path.splitext(media)[1].lstrip(".").lower()
        if ext not in ("png", "jpg", "jpeg"):
            continue
        nb = columbia_bytes(modelo, "png" if ext == "png" else "jpeg")
        if nb:
            media_new[media] = nb; n_ok += 1
        else:
            n_miss += 1
    print(f"{name}: fotos reemplazadas={n_ok} sin_foto={n_miss}", flush=True)
z.close()

# 3) reescribir el .xlsx cambiando SOLO esos media
with zipfile.ZipFile(SRC) as zin, \
        zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = media_new.get(item.filename, zin.read(item.filename))
        zout.writestr(item, data)

print(f"\nLISTO: {n_ok} fotos reemplazadas, {n_miss} sin foto -> {OUT}", flush=True)
