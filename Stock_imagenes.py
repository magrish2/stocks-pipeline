"""
STOCK CON IMÁGENES - Con agrupación por Modelo-Color
=====================================================
Uso CLI:
    python stock_imagenes.py archivo.xlsx
    python stock_imagenes.py archivo.xlsx --ref ref1.xlsx --ref ref2.xlsx

Uso GUI:
    python stock_imagenes.py

Requisitos:
    pip install openpyxl aiohttp pillow

Fuentes de imagen (en orden de prioridad):
    1. CDN oficial de crocs.com  → construido directo desde el SKU (900x900, fondo transparente)
    2. Archivo(s) de referencia  → imágenes de otros stocks con --ref
    3. Caché en disco            → _img_cache/ (no re-descarga lo que ya tiene)
"""

import sys
import re
import asyncio
import io
import os
import tempfile
import zipfile
import subprocess
import threading
import aiohttp
import xml.etree.ElementTree as ET
from urllib.parse import quote as url_quote
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

IMG_WIDTH_PX   = 220
IMG_HEIGHT_PX  = 180
ROW_HEIGHT_PTS = 140
TIMEOUT_SEC    = 6
MAX_CONCURRENT = 50
CACHE_DIR      = Path("_img_cache")
REGISTRO_DIR   = Path("_registro")
CDN_BASE       = "https://media.crocs.com/images/f_auto%2Cq_auto%2Cw_900%2Ch_900%2Cc_pad%2Cb_transparent/products"
SHOPIFY_CDN    = "https://cdn.shopify.com/s/files/1/0937/6150/3518/files"

_shopify_catalog: dict = {}

# ─── NAMESPACES DRAWING XML ───────────────────────────────────────────────────

NS_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
NS_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

# ─── PATRONES DE COLUMNAS ─────────────────────────────────────────────────────

SKU_PAT  = re.compile(r'^(sku|cod|c[oó]digo|ref|referencia)$', re.IGNORECASE)
DESC_PAT = re.compile(r'^(desc|descripci[oó]n|nombre|producto|art[ií]culo|item|marca)$', re.IGNORECASE)
IMG_PAT  = re.compile(r'^(imagen|foto|image|url.?imagen|url.?foto)$', re.IGNORECASE)
MC_PAT   = re.compile(r'^(modelo.?color|modelocolor|modcolor|mc|model.?color|mod.?col)$', re.IGNORECASE)

REQ_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

print_lock = threading.Lock()

# ─── DESCARGA Y CACHÉ (ASYNC) ─────────────────────────────────────────────────

async def download_image_async(session, url, cache_key):
    _cache = CACHE_DIR
    _cache.mkdir(exist_ok=True)
    safe = re.sub(r'[^\w]', '_', cache_key)[:70]
    cache_file = _cache / f"{safe}.png"

    if cache_file.exists():
        return cache_file.read_bytes()

    try:
        timeout = aiohttp.ClientTimeout(total=TIMEOUT_SEC)
        async with session.get(url, timeout=timeout) as resp:
            if resp.status != 200:
                return None
            raw = await resp.read()

        img = PILImage.open(io.BytesIO(raw))
        if img.mode not in ('RGB', 'RGBA'):
            img = img.convert('RGBA' if 'transparency' in img.info else 'RGB')
        img.thumbnail((IMG_WIDTH_PX, IMG_HEIGHT_PX), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        data = buf.getvalue()
        cache_file.write_bytes(data)
        return data
    except Exception:
        return None


async def vtex_image_bytes_async(session, domain, sku, cache_key):
    try:
        api_url = (
            f"https://{domain}/api/catalog_system/pub/products/search"
            f"?q={url_quote(sku)}&sc=1&_from=0&_to=1"
        )
        timeout = aiohttp.ClientTimeout(total=TIMEOUT_SEC)
        async with session.get(api_url, timeout=timeout) as resp:
            if resp.status != 200:
                return None, None
            products = await resp.json(content_type=None)
        if not products:
            return None, None
        images = products[0].get('items', [{}])[0].get('images', [])
        img_url = images[0].get('imageUrl', '') if images else ''
        if not img_url:
            return None, None
        data = await download_image_async(session, img_url, cache_key)
        return data, img_url
    except Exception:
        return None, None


async def build_shopify_catalog_async(session, log_fn=print):
    global _shopify_catalog
    catalog = {}
    page = 1
    log_fn("  Cargando catálogo de crocs.com.ar...")
    while True:
        url = f"https://www.crocs.com.ar/collections/all/products.json?limit=250&page={page}"
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                if resp.status != 200:
                    break
                data = await resp.json(content_type=None)
        except Exception:
            break
        products = data.get('products', [])
        if not products:
            break
        for product in products:
            images = product.get('images', [])
            fallback_url = images[0].get('src', '') if images else ''
            img_by_id = {img['id']: img.get('src', '') for img in images if img.get('id')}
            for variant in product.get('variants', []):
                sku_v = str(variant.get('sku') or '').strip().upper()
                if not sku_v:
                    continue
                fi = variant.get('featured_image') or {}
                img_url = fi.get('src', '') or img_by_id.get(fi.get('id', ''), '') or fallback_url
                if not img_url:
                    continue
                # Indexar por SKU completo (con talle: C10006-C0DA-C12/13)
                if sku_v not in catalog:
                    catalog[sku_v] = img_url
                # Indexar también por prefijo MC sin talle (C10006-C0DA)
                # para cuando el Excel usa solo modelo-color
                mc_match = re.match(r'^(C\d+-C\w+)-\S', sku_v)
                if mc_match:
                    mc_key = mc_match.group(1)
                    if mc_key not in catalog:
                        catalog[mc_key] = img_url
        page += 1
    _shopify_catalog = catalog
    log_fn(f"  {len(catalog)} productos indexados en crocs.com.ar")


async def crocs_cdn_bytes_async(session, sku, cache_key):
    candidates = []
    sku_up = sku.upper()
    for n in ['1', '2', '3']:
        candidates.append((f"{SHOPIFY_CDN}/{sku_up}-{n}.jpg", 'crocs.com.ar'))
    for n in ['1', '2', '3']:
        candidates.append((f"https://www.crocs.com.ar/cdn/shop/files/{sku_up}-{n}.jpg", 'crocs.com.ar'))

    m = re.match(r'C(\d+)-C(\w+)', sku_up)
    if m:
        prod, color = m.group(1), m.group(2)
        for suf in ['_ALT100', '_ALT1', '']:
            for fname in ['crocs.jpg', 'jibbitz.jpg']:
                url = f"{CDN_BASE}/{prod}_{color}{suf}/{fname}"
                candidates.append((url, 'crocs.com'))

    async def try_url(url, src):
        data = await download_image_async(session, url, cache_key)
        return (data, src, url) if data else (None, None, None)

    async def try_vtex():
        data, img_url = await vtex_image_bytes_async(session, 'www.megasports.com.ar', sku, cache_key)
        return (data, 'megasports', img_url) if data else (None, None, None)

    async def try_catalog():
        img_url = _shopify_catalog.get(sku_up, '')
        if not img_url:
            return None, None, None
        data = await download_image_async(session, img_url, cache_key)
        return (data, 'crocs.com.ar', img_url) if data else (None, None, None)

    tasks = [asyncio.create_task(try_url(url, src)) for url, src in candidates]
    tasks.append(asyncio.create_task(try_vtex()))
    tasks.append(asyncio.create_task(try_catalog()))

    pending = set(tasks)
    while pending:
        done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
        for task in done:
            try:
                data, src, url = task.result()
                if data:
                    for t in pending:
                        t.cancel()
                    return data, src, url
            except Exception:
                pass
    return None, None, None


async def resolve_image_async(session, mc, sku, desc, ref_images):
    safe = re.sub(r'[^\w]', '_', mc)[:70]
    cache_file = CACHE_DIR / f"{safe}.png"  # CACHE_DIR global, actualizado antes de llamar
    if cache_file.exists():
        return cache_file.read_bytes(), 'caché', None
    # Usar mc (ej: "C207010-C060") para CDN, no el SKU completo con talles
    data, src, url = await crocs_cdn_bytes_async(session, mc, mc)
    if data:
        return data, src, url
    if desc.lower() in ref_images:
        return ref_images[desc.lower()], 'referencia', None
    return None, None, None

def emf_to_png_bytes(emf_data):
    if sys.platform != 'win32':
        return None
    emf_path = png_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.emf', delete=False) as f:
            f.write(emf_data)
            emf_path = f.name
        png_path = emf_path[:-4] + '.png'
        emf_ps = emf_path.replace('\\', '/')
        png_ps = png_path.replace('\\', '/')
        ps = f"""
Add-Type -AssemblyName System.Drawing
$img = [System.Drawing.Image]::FromFile('{emf_ps}')
$w = [Math]::Max($img.Width, 200); $h = [Math]::Max($img.Height, 200)
$bmp = New-Object System.Drawing.Bitmap($w, $h)
$g = [System.Drawing.Graphics]::FromImage($bmp)
$g.Clear([System.Drawing.Color]::White)
$g.DrawImage($img, 0, 0, $w, $h)
$g.Dispose()
$bmp.Save('{png_ps}', [System.Drawing.Imaging.ImageFormat]::Png)
$bmp.Dispose(); $img.Dispose()
"""
        subprocess.run(['powershell', '-NoProfile', '-Command', ps],
                       capture_output=True, timeout=30)
        if os.path.exists(png_path):
            with open(png_path, 'rb') as f:
                return f.read()
    except Exception:
        pass
    finally:
        for p in [emf_path, png_path]:
            if p and os.path.exists(p):
                try: os.unlink(p)
                except: pass
    return None

def load_reference_images(ref_paths, log_fn=print):
    merged = {}
    for ref_path in ref_paths:
        log_fn(f"\n── Referencia: {ref_path} ──")
        result = _load_single_ref(ref_path, log_fn)
        new = {k: v for k, v in result.items() if k not in merged}
        merged.update(new)
        log_fn(f"  Aportadas: {len(new)} (total acumulado: {len(merged)})")
    return merged

def _load_single_ref(ref_path, log_fn=print):
    with zipfile.ZipFile(ref_path) as z:
        names = z.namelist()
        if 'xl/drawings/_rels/drawing1.xml.rels' not in names:
            log_fn("  ⚠ Sin imágenes en drawing1.")
            return {}
        rels_tree = ET.fromstring(z.read('xl/drawings/_rels/drawing1.xml.rels'))
        rid_to_file = {
            rel.attrib.get('Id'): rel.attrib.get('Target', '').replace('../media/', '')
            for rel in rels_tree
        }
        drawing_tree = ET.fromstring(z.read('xl/drawings/drawing1.xml'))
        row_to_imgfile = {}
        for tag in (f'{{{NS_XDR}}}twoCellAnchor', f'{{{NS_XDR}}}oneCellAnchor'):
            for anchor in drawing_tree.findall(tag):
                from_el = anchor.find(f'{{{NS_XDR}}}from')
                if from_el is None: continue
                row_el = from_el.find(f'{{{NS_XDR}}}row')
                if row_el is None: continue
                blip = anchor.find(f'.//{{{NS_A}}}blip')
                if blip is None: continue
                rid = blip.attrib.get(f'{{{NS_R}}}embed')
                if rid:
                    row_to_imgfile[int(row_el.text)] = rid_to_file.get(rid, '')
        unique_files = set(row_to_imgfile.values()) - {''}
        raw_images = {
            fname: z.read(f'xl/media/{fname}')
            for fname in unique_files
            if f'xl/media/{fname}' in names
        }
    converted = {}
    for fname, raw in raw_images.items():
        if fname.lower().endswith('.emf'):
            converted[fname] = emf_to_png_bytes(raw)
        else:
            try:
                img = PILImage.open(io.BytesIO(raw))
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                buf = io.BytesIO()
                img.save(buf, 'PNG')
                converted[fname] = buf.getvalue()
            except Exception:
                converted[fname] = None
    wb_ref = load_workbook(ref_path, read_only=True, data_only=True)
    ws_ref = wb_ref.active
    ref_data = list(ws_ref.values)
    wb_ref.close()
    if not ref_data: return {}
    ref_cols = detect_columns(list(ref_data[0]))
    if ref_cols['desc'] is None: return {}
    desc_to_png = {}
    last_fname = ''
    for row_i, row in enumerate(ref_data[1:]):
        excel_row = row_i + 1
        fname = row_to_imgfile.get(excel_row, '') or last_fname
        if fname: last_fname = fname
        png = converted.get(fname)
        if not png: continue
        desc = str(row[ref_cols['desc']] or '').strip().lower()
        if desc and desc not in desc_to_png:
            desc_to_png[desc] = png
    return desc_to_png

def detect_columns(headers):
    cols = {'sku': None, 'desc': None, 'img': None, 'mc': None}
    for i, h in enumerate(headers):
        s = str(h or '').strip()
        if cols['mc']   is None and MC_PAT.match(s):   cols['mc']   = i
        if cols['sku']  is None and SKU_PAT.match(s):  cols['sku']  = i
        if cols['desc'] is None and DESC_PAT.match(s): cols['desc'] = i
        if cols['img']  is None and IMG_PAT.match(s):  cols['img']  = i
    return cols

def ask_column(headers, label):
    print(f"\n  No se detectó '{label}'. Columnas disponibles:")
    for i, h in enumerate(headers):
        print(f"    [{i:2d}] {h}")
    while True:
        try:
            val = int(input(f"  Número de columna para {label}: "))
            if 0 <= val < len(headers):
                return val
        except ValueError:
            pass


def ask_column_gui(root, headers, label):
    """Muestra un diálogo modal en el hilo principal y devuelve el índice elegido."""
    import tkinter as tk
    from tkinter import ttk

    BG_FRAME = "#a00000"
    BG_ENTRY = "#b31010"
    FG       = "#ffe8e8"
    FG_HINT  = "#ffaaaa"
    BG_BTN   = "#cc2222"
    FG_BTN   = "#ffffff"

    result = [None]
    event  = threading.Event()

    def show():
        dlg = tk.Toplevel(root)
        dlg.title("Seleccionar columna")
        dlg.configure(bg=BG_FRAME)
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(dlg, text=f"No se detectó automáticamente: \"{label}\"",
                 bg=BG_FRAME, fg=FG, font=("Helvetica", 11, "bold")).pack(padx=20, pady=(16, 4))
        tk.Label(dlg, text="Seleccioná la columna que corresponde:",
                 bg=BG_FRAME, fg=FG_HINT).pack(padx=20)

        lb_frame = tk.Frame(dlg, bg=BG_FRAME)
        lb_frame.pack(padx=20, pady=10, fill='both', expand=True)

        sb = ttk.Scrollbar(lb_frame, orient='vertical')
        lb = tk.Listbox(lb_frame, yscrollcommand=sb.set, height=min(len(headers), 12),
                        bg=BG_ENTRY, fg=FG, selectbackground=BG_BTN, selectforeground=FG_BTN,
                        relief="flat", borderwidth=1, highlightthickness=1,
                        highlightbackground=BG_BTN, font=("Courier", 10))
        sb.config(command=lb.yview)
        lb.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        for i, h in enumerate(headers):
            lb.insert('end', f"[{i:2d}]  {h}")
        lb.selection_set(0)

        def on_ok():
            sel = lb.curselection()
            if not sel:
                return
            result[0] = sel[0]
            dlg.destroy()
            event.set()

        def on_close():
            event.set()

        dlg.protocol("WM_DELETE_WINDOW", on_close)
        ttk.Button(dlg, text="Confirmar", command=on_ok).pack(pady=(0, 16))

        dlg.update_idletasks()
        w, h = dlg.winfo_reqwidth(), dlg.winfo_reqheight()
        x = root.winfo_x() + (root.winfo_width()  - w) // 2
        y = root.winfo_y() + (root.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

    root.after(0, show)
    event.wait()
    return result[0]

# ─── ESTILOS ──────────────────────────────────────────────────────────────────

_HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_HDR_SIDE   = Side(style='thin', color="BBBBBB")
_HDR_BRD    = Border(left=_HDR_SIDE, right=_HDR_SIDE, top=_HDR_SIDE, bottom=_HDR_SIDE)
_HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ROW_SIDE       = Side(style='thin', color="DDDDDD")
_ROW_BRD        = Border(left=_ROW_SIDE, right=_ROW_SIDE, top=_ROW_SIDE, bottom=_ROW_SIDE)
_ROW_FONT       = Font(name="Arial", size=9)
_ROW_ALIGN      = Alignment(vertical="center", wrap_text=True)
_ROW_FILL_GROUP = PatternFill("solid", fgColor="EFF4FB")
_ROW_FILL_EVEN  = PatternFill("solid", fgColor="F7F9FC")
_ROW_FILL_ODD   = PatternFill("solid", fgColor="FFFFFF")

def style_header(ws, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT; cell.border = _HDR_BRD
        cell.alignment = _HDR_ALIGN
    ws.row_dimensions[1].height = 26

def style_row(ws, row_num, n_cols, is_same_group):
    fill = _ROW_FILL_GROUP if is_same_group else (_ROW_FILL_EVEN if row_num % 2 == 0 else _ROW_FILL_ODD)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = _ROW_FONT
        cell.alignment = _ROW_ALIGN
        cell.border = _ROW_BRD
        cell.fill = fill

# ─── PROCESO PRINCIPAL ────────────────────────────────────────────────────────

def process(input_path, ref_paths, log_fn=print, ask_column_fn=None):
    log_fn(f"\n{'='*60}")
    log_fn("  STOCK CON IMÁGENES — agrupación por Modelo-Color")
    log_fn(f"{'='*60}")
    log_fn(f"  Archivo : {input_path}")

    ref_images = load_reference_images(ref_paths, log_fn) if ref_paths else {}
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    all_data = list(ws.values)
    if not all_data:
        log_fn("❌ El archivo está vacío.")
        return None

    headers_raw = list(all_data[0])
    rows = [list(r) for r in all_data[1:] if any(c not in (None, '') for c in r)]
    n_total = len(rows)
    log_fn(f"  Filas de datos: {n_total}")

    cols = detect_columns(headers_raw)
    _ask = ask_column_fn if ask_column_fn else ask_column
    if cols['sku']  is None:
        log_fn(f"⚠️  No se detectó la columna SKU/Código — seleccionala manualmente.")
        cols['sku'] = _ask(headers_raw, "SKU / Código")
        if cols['sku'] is None:
            log_fn("❌ No se seleccionó columna SKU. Se cancela el proceso.")
            return None
    if cols['desc'] is None:
        log_fn(f"⚠️  No se detectó la columna Descripción — seleccionala manualmente.")
        cols['desc'] = _ask(headers_raw, "Descripción")
        if cols['desc'] is None:
            log_fn("❌ No se seleccionó columna Descripción. Se cancela el proceso.")
            return None
    if cols['mc'] is None:
        log_fn(f"⚠️  No se detectó la columna Modelo-Color — seleccionala manualmente (o cancelá para omitirla).")
        cols['mc'] = _ask(headers_raw, "Modelo-Color (opcional)")
    if cols['img'] is None:
        log_fn(f"⚠️  No se detectó la columna Imagen — seleccionala manualmente (o cancelá para crearla automáticamente).")
        cols['img'] = _ask(headers_raw, "Imagen / URL Imagen (opcional)")

    use_mc = cols['mc'] is not None
    if cols['img'] is None:
        # El usuario eligió no mapear ninguna columna → crear columna nueva al final
        headers_raw.append("Imagen")
        cols['img'] = len(headers_raw) - 1
        for row in rows: row.append(None)
        log_fn("  → Se creará una columna 'Imagen' nueva al final del archivo.")

    n_cols = len(headers_raw)
    mc_cache, mc_source, mc_url = {}, {}, {}

    # CACHE_DIR relativo al archivo de entrada para garantizar escritura
    global CACHE_DIR
    CACHE_DIR = Path(input_path).parent / "_img_cache"
    CACHE_DIR.mkdir(exist_ok=True)
    log_fn(f"  Caché en: {CACHE_DIR}")

    if use_mc:
        mc_repr = {}
        for row in rows:
            mc = str(row[cols['mc']] or '').strip()
            if mc and mc not in mc_repr:
                mc_repr[mc] = (str(row[cols['sku']] or '').strip(), str(row[cols['desc']] or '').strip())

        log_fn(f"\n📥 Descargando imágenes para {len(mc_repr)} modelos únicos...")
        sample = list(mc_repr.items())[:5]
        log_fn(f"  Muestra de SKUs a buscar:")
        for mc, (sku, desc) in sample:
            log_fn(f"    MC='{mc}'  SKU='{sku}'  DESC='{desc}'")

        async def resolve_all():
            sem = asyncio.Semaphore(MAX_CONCURRENT)
            connector = aiohttp.TCPConnector(limit=MAX_CONCURRENT)
            async with aiohttp.ClientSession(headers=REQ_HEADERS, connector=connector) as session:
                await build_shopify_catalog_async(session, log_fn)
                if _shopify_catalog:
                    sample_keys = list(_shopify_catalog.keys())[:5]
                    log_fn(f"  Muestra de SKUs en catálogo: {sample_keys}")
                    matches = sum(1 for _, (sku, _) in mc_repr.items() if sku.upper() in _shopify_catalog)
                    log_fn(f"  Coincidencias SKU↔catálogo: {matches}/{len(mc_repr)}")
                async def bounded(mc, sku, desc):
                    async with sem:
                        data, src, url = await resolve_image_async(session, mc, sku, desc, ref_images)
                        return mc, data, src, url
                coros = [bounded(mc, sku, desc) for mc, (sku, desc) in mc_repr.items()]
                done_count = 0
                for coro in asyncio.as_completed(coros):
                    mc, data, src, url = await coro
                    mc_cache[mc], mc_source[mc], mc_url[mc] = data, src, url
                    done_count += 1
                    if done_count % 10 == 0 or done_count == len(mc_repr):
                        found_so_far = sum(1 for v in mc_cache.values() if v)
                        log_fn(f"  {done_count}/{len(mc_repr)} procesados — {found_so_far} con imagen")
        asyncio.run(resolve_all())

    # ── Fase 2: Escribir Excel ──
    log_fn(f"\n✍️  Generando Excel con imágenes...")
    ws.delete_rows(1, ws.max_row)
    ws.append(headers_raw)
    style_header(ws, n_cols)

    for i, _ in enumerate(headers_raw):
        cl = get_column_letter(i + 1)
        if   i == cols['img']:  ws.column_dimensions[cl].width = 30
        elif i == cols['desc']: ws.column_dimensions[cl].width = 32
        elif i == cols['mc']:   ws.column_dimensions[cl].width = 20
        else:                   ws.column_dimensions[cl].width = 16
    ws.freeze_panes = "A2"

    found = failed = 0
    prev_mc = None
    for i, row in enumerate(rows):
        row_num = i + 2
        mc_val = str(row[cols['mc']] or '').strip() if use_mc else None
        sku = str(row[cols['sku']] or '').strip()
        desc = str(row[cols['desc']] or '').strip()
        is_same_group = (mc_val is not None and mc_val == prev_mc)

        ws.append(row)
        style_row(ws, row_num, n_cols, is_same_group)
        ws.row_dimensions[row_num].height = ROW_HEIGHT_PTS
        prev_mc = mc_val

        img_cell = ws.cell(row=row_num, column=cols['img'] + 1)
        img_cell.alignment = Alignment(horizontal="center", vertical="center")

        img_data = mc_cache.get(mc_val) if use_mc else None
        if not img_data:
            safe = re.sub(r'[^\w]', '_', mc_val or sku)[:70]
            cached = CACHE_DIR / f"{safe}.png"
            if cached.exists():
                img_data = cached.read_bytes()
        if img_data:
            xl_img = XLImage(io.BytesIO(img_data))
            anchor = TwoCellAnchor()
            anchor.editAs = 'twoCell'
            anchor._from = AnchorMarker(col=cols['img'],     colOff=0, row=row_num - 1, rowOff=0)
            anchor.to    = AnchorMarker(col=cols['img'] + 1, colOff=0, row=row_num,     rowOff=0)
            xl_img.anchor = anchor
            ws.add_image(xl_img)
            img_cell.value = None
            found += 1
        else:
            img_cell.value = "Sin imagen"
            failed += 1

    out_path = re.sub(r'\.xlsx?$', '_con_imagenes.xlsx', input_path)
    wb.save(out_path)
    _lock_images_in_xlsx(out_path, log_fn)
    log_fn(f"\n✅ Generado: {out_path}")
    log_fn(f"   Con imagen: {found} | Sin imagen: {failed}")
    return out_path


def _lock_images_in_xlsx(path, log_fn=print):
    """Post-procesa el xlsx para bloquear todas las imágenes en sus celdas:
    noSelect + noMove + noResize sobre cada <a:picLocks> del drawing XML."""
    tmp = path + '.lck.tmp'

    def patch_drawing(text):
        # openpyxl serializa sin prefix: <cNvPicPr /> o <cNvPicPr>...</cNvPicPr>
        LOCKS = '<a:picLocks/>'

        def replace_block(m):
            inner = m.group(1)
            inner = re.sub(r'<a:picLocks[^/]*/>', '', inner)
            inner = re.sub(r'<a:picLocks[^>]*>.*?</a:picLocks>', '', inner, flags=re.DOTALL)
            inner += LOCKS
            return f'<cNvPicPr>{inner}</cNvPicPr>'

        # self-closing → expandir y agregar locks
        text = text.replace('<cNvPicPr />', f'<cNvPicPr>{LOCKS}</cNvPicPr>')
        text = text.replace('<cNvPicPr/>', f'<cNvPicPr>{LOCKS}</cNvPicPr>')
        # con contenido → reemplazar/agregar locks
        text = re.sub(r'<cNvPicPr>(.*?)</cNvPicPr>', replace_block, text, flags=re.DOTALL)
        return text

    with zipfile.ZipFile(path, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if re.match(r'xl/drawings/drawing\d+\.xml$', item.filename):
                    data = patch_drawing(data.decode('utf-8')).encode('utf-8')
                zout.writestr(item, data)

    os.replace(tmp, path)
    log_fn("  🔒 Imágenes ancladas y bloqueadas en celda")


# ─── INTERFAZ GRÁFICA ─────────────────────────────────────────────────────────

def main():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    BG        = "#8b0000"
    BG_FRAME  = "#a00000"
    BG_ENTRY  = "#b31010"
    FG        = "#ffe8e8"
    FG_HINT   = "#ffaaaa"
    BG_LOG    = "#5c0000"
    FG_LOG    = "#ffcccc"
    BG_BTN    = "#cc2222"
    FG_BTN    = "#ffffff"

    root = tk.Tk()
    root.title("Stock con Imágenes — Crocs")
    root.resizable(True, True)
    root.configure(bg=BG)

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure(".",          background=BG_FRAME, foreground=FG, fieldbackground=BG_ENTRY)
    style.configure("TFrame",     background=BG_FRAME)
    style.configure("TLabel",     background=BG_FRAME, foreground=FG)
    style.configure("TEntry",     fieldbackground=BG_ENTRY, foreground=FG, insertcolor=FG,
                                  bordercolor="#2d5be3", lightcolor=BG_ENTRY, darkcolor=BG_ENTRY)
    style.configure("TButton",    background=BG_BTN, foreground=FG_BTN, bordercolor=BG_BTN, focuscolor=BG_BTN)
    style.map("TButton",          background=[("active", "#3d6ef5"), ("disabled", "#1e2f54")],
                                  foreground=[("disabled", FG_HINT)])
    style.configure("TScrollbar", background=BG_ENTRY, troughcolor=BG_LOG, bordercolor=BG_FRAME, arrowcolor=FG_HINT)

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill='both', expand=True)
    frame.columnconfigure(1, weight=1)
    frame.rowconfigure(6, weight=1)

    ttk.Label(frame, text="Stock con Imágenes — Crocs",
              font=("Helvetica", 14, "bold"), foreground="#7eb3ff").grid(
        row=0, column=0, columnspan=3, pady=(0, 20))

    # ── Cola de archivos Excel ────────────────────────────────────────────────
    ttk.Label(frame, text="Archivos Excel:").grid(row=1, column=0, sticky='nw')

    queue_frame = ttk.Frame(frame)
    queue_frame.grid(row=1, column=1, columnspan=2, sticky='ew')
    queue_frame.columnconfigure(0, weight=1)

    queue_lb = tk.Listbox(queue_frame, height=4, bg=BG_ENTRY, fg=FG,
                          selectbackground="#2d5be3", selectforeground=FG_BTN,
                          relief="flat", borderwidth=1, highlightthickness=1,
                          highlightbackground="#2d5be3", highlightcolor="#3d6ef5")
    queue_lb.grid(row=0, column=0, sticky='ew')

    queue_scroll = ttk.Scrollbar(queue_frame, orient='vertical', command=queue_lb.yview)
    queue_scroll.grid(row=0, column=1, sticky='ns')
    queue_lb.configure(yscrollcommand=queue_scroll.set)

    queue_btn_frame = ttk.Frame(queue_frame)
    queue_btn_frame.grid(row=1, column=0, sticky='w', pady=(4, 0))

    def add_excel():
        paths = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx *.xls")])
        for p in paths:
            if p not in queue_lb.get(0, 'end'):
                queue_lb.insert('end', p)

    def remove_excel():
        for i in reversed(queue_lb.curselection()):
            queue_lb.delete(i)

    def move_up():
        sel = queue_lb.curselection()
        if not sel or sel[0] == 0:
            return
        i = sel[0]
        val = queue_lb.get(i)
        queue_lb.delete(i)
        queue_lb.insert(i - 1, val)
        queue_lb.selection_set(i - 1)

    def move_down():
        sel = queue_lb.curselection()
        if not sel or sel[0] == queue_lb.size() - 1:
            return
        i = sel[0]
        val = queue_lb.get(i)
        queue_lb.delete(i)
        queue_lb.insert(i + 1, val)
        queue_lb.selection_set(i + 1)

    ttk.Button(queue_btn_frame, text="+ Agregar", command=add_excel).pack(side='left', padx=(0, 6))
    ttk.Button(queue_btn_frame, text="− Quitar",  command=remove_excel).pack(side='left', padx=(0, 6))
    ttk.Button(queue_btn_frame, text="↑",         command=move_up,      width=3).pack(side='left', padx=(0, 4))
    ttk.Button(queue_btn_frame, text="↓",         command=move_down,    width=3).pack(side='left')
    ttk.Label(queue_frame, text="Se procesan en orden de arriba hacia abajo",
              foreground=FG_HINT, font=("TkDefaultFont", 8)).grid(row=2, column=0, sticky='w', pady=(2, 0))

    # ── Archivos de referencia ────────────────────────────────────────────────
    ttk.Label(frame, text="Refs (opcional):").grid(row=3, column=0, sticky='nw', pady=(12, 0))

    ref_frame = ttk.Frame(frame)
    ref_frame.grid(row=3, column=1, columnspan=2, sticky='ew', pady=(12, 0))
    ref_frame.columnconfigure(0, weight=1)

    ref_lb = tk.Listbox(ref_frame, height=3, bg=BG_ENTRY, fg=FG,
                        selectbackground="#2d5be3", selectforeground=FG_BTN,
                        relief="flat", borderwidth=1, highlightthickness=1,
                        highlightbackground="#2d5be3", highlightcolor="#3d6ef5")
    ref_lb.grid(row=0, column=0, sticky='ew')

    ref_scroll = ttk.Scrollbar(ref_frame, orient='vertical', command=ref_lb.yview)
    ref_scroll.grid(row=0, column=1, sticky='ns')
    ref_lb.configure(yscrollcommand=ref_scroll.set)

    ref_btn_frame = ttk.Frame(ref_frame)
    ref_btn_frame.grid(row=1, column=0, sticky='w', pady=(4, 0))

    def add_ref():
        paths = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx *.xls")])
        for p in paths:
            if p not in ref_lb.get(0, 'end'):
                ref_lb.insert('end', p)

    def remove_ref():
        for i in reversed(ref_lb.curselection()):
            ref_lb.delete(i)

    ttk.Button(ref_btn_frame, text="+ Agregar", command=add_ref).pack(side='left', padx=(0, 6))
    ttk.Button(ref_btn_frame, text="− Quitar",  command=remove_ref).pack(side='left')
    ttk.Label(ref_frame, text="Excels con imágenes existentes para reutilizar",
              foreground=FG_HINT, font=("TkDefaultFont", 8)).grid(row=2, column=0, sticky='w', pady=(2, 0))

    # ── Log ───────────────────────────────────────────────────────────────────
    ttk.Label(frame, text="Log de ejecución:").grid(row=5, column=0, sticky='nw', pady=(16, 0))
    log_text = tk.Text(frame, width=70, height=16, state='disabled', font=("Courier", 9),
                       bg=BG_LOG, fg=FG_LOG, insertbackground=FG_LOG,
                       relief="flat", bd=1, highlightthickness=1,
                       highlightbackground="#2d5be3", highlightcolor="#3d6ef5")
    log_text.grid(row=6, column=0, columnspan=3, pady=(4, 12), sticky='nsew')

    scrollbar = ttk.Scrollbar(frame, orient='vertical', command=log_text.yview)
    scrollbar.grid(row=6, column=3, pady=(4, 12), sticky='ns')
    log_text.configure(yscrollcommand=scrollbar.set)

    def log(msg):
        log_text.configure(state='normal')
        log_text.insert('end', msg + '\n')
        log_text.see('end')
        log_text.configure(state='disabled')

    # ── Botón procesar ────────────────────────────────────────────────────────
    def iniciar():
        rutas = list(queue_lb.get(0, 'end'))
        if not rutas:
            messagebox.showwarning("Falta dato", "Agregá al menos un archivo Excel.")
            return

        refs = list(ref_lb.get(0, 'end'))

        btn.configure(state='disabled')
        log_text.configure(state='normal')
        log_text.delete("1.0", "end")
        log_text.configure(state='disabled')

        def run():
            try:
                ask_fn = lambda headers, label: ask_column_gui(root, headers, label)
                total = len(rutas)
                for idx, ruta in enumerate(rutas, 1):
                    log(f"\n{'─'*50}")
                    log(f"📄 Archivo {idx}/{total}: {Path(ruta).name}")
                    try:
                        out = process(ruta, refs, log_fn=log, ask_column_fn=ask_fn)
                        if out:
                            log(f"📂 Guardado en: {out}")
                    except Exception as ex:
                        import traceback
                        log(f"❌ Error en {Path(ruta).name}: {ex}")
                        log(traceback.format_exc())
                log(f"\n{'='*50}")
                log(f"✅ Todos los archivos procesados ({total}/{total})")
            finally:
                root.after(0, lambda: btn.configure(state='normal'))

        threading.Thread(target=run, daemon=True).start()

    btn = ttk.Button(frame, text="▶  Generar Stock con Imágenes", command=iniciar)
    btn.grid(row=7, column=0, columnspan=3, pady=(0, 8))

    root.update_idletasks()
    w = root.winfo_reqwidth()
    h = root.winfo_reqheight()
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")
    root.minsize(w, h)

    root.mainloop()


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        main()
    else:
        ref_paths, input_args = [], []
        i = 0
        while i < len(args):
            if args[i] == '--ref': ref_paths.append(args[i+1]); i += 2
            else: input_args.append(args[i]); i += 1
        process(" ".join(input_args), ref_paths)
