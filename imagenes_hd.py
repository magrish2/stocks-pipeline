#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mejora de calidad de imágenes con IA (Real-ESRGAN local) para Excels y artículos.

Dos modos:

  1) excel  — Toma cualquier .xlsx y deja sus fotos en HD (super-resolución IA),
              conservando posiciones y todo el resto. Genera '<archivo> HD.xlsx'.

        python imagenes_hd.py excel --input "MI ARCHIVO.xlsx"
        python imagenes_hd.py excel --input "*.xlsx"        # varios

  2) articulos — Le das códigos/SKU y descarga las imágenes en alta calidad
                 (web oficial Reebok) y opcionalmente las mejora con IA.

        python imagenes_hd.py articulos RBK1100074190-8 RBK1100225472
        python imagenes_hd.py articulos --from-excel "stock.xlsx" --col "Número"
        python imagenes_hd.py articulos --from-txt codigos.txt

El motor IA es Real-ESRGAN (binario ncnn en ./tools, offline, gratis).
"""

from __future__ import annotations

import argparse
import glob
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from io import BytesIO

import requests
from PIL import Image

# Reutilizamos la lógica de resolución de imágenes Reebok del otro script.
import normalizar_stock as N

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.join(BASE_DIR, "tools")


def _find_engine():
    """Ubica el binario Real-ESRGAN (Windows .exe o macOS/Linux), en tools/."""
    candidates = [
        os.path.join(TOOLS_DIR, "realesrgan-ncnn-vulkan.exe"),  # Windows
        os.path.join(TOOLS_DIR, "realesrgan-ncnn-vulkan"),      # macOS / Linux
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return candidates[0] if sys.platform.startswith("win") else candidates[1]


REALESRGAN_BIN = _find_engine()
REALESRGAN_MODELS = os.path.join(TOOLS_DIR, "models")
DEFAULT_MODEL = "realesrgan-x4plus"   # foto realista (no anime)

IMG_EXTS = ("png", "jpg", "jpeg", "webp")


# --------------------------------------------------------------------------- #
# Motor IA
# --------------------------------------------------------------------------- #
def have_engine():
    return os.path.exists(REALESRGAN_BIN)


def ai_upscale(in_path, out_path, model=DEFAULT_MODEL):
    """Corre Real-ESRGAN sobre un archivo. Devuelve True si generó la salida."""
    try:
        subprocess.run(
            [REALESRGAN_BIN, "-i", in_path, "-o", out_path,
             "-s", "4", "-n", model, "-m", REALESRGAN_MODELS],
            capture_output=True, timeout=120,
        )
    except (subprocess.SubprocessError, OSError):
        return False
    return os.path.exists(out_path)


def enhance_bytes(raw, ext, max_px, model=DEFAULT_MODEL, min_px=900):
    """
    Mejora una imagen (bytes) con IA y la limita a max_px. Devuelve bytes (mismo
    formato) o None si no hace falta / falla. Salta las que ya son grandes.
    """
    try:
        im = Image.open(BytesIO(raw))
        w, h = im.size
    except Exception:
        return None
    if max(w, h) >= min_px:
        return None  # ya tiene buena resolución, no la tocamos

    tmpd = tempfile.mkdtemp()
    try:
        inp = os.path.join(tmpd, "in." + (ext if ext in IMG_EXTS else "png"))
        outp = os.path.join(tmpd, "out.png")
        with open(inp, "wb") as fh:
            fh.write(raw)
        if not ai_upscale(inp, outp, model=model):
            return None
        hd = Image.open(outp)
        if max(hd.size) > max_px:
            hd.thumbnail((max_px, max_px), Image.LANCZOS)
        buf = BytesIO()
        if ext in ("jpg", "jpeg"):
            hd.convert("RGB").save(buf, "JPEG", quality=92)
        elif ext == "webp":
            hd.save(buf, "WEBP", quality=92)
        else:
            hd.save(buf, "PNG")
        return buf.getvalue()
    finally:
        shutil.rmtree(tmpd, ignore_errors=True)


# --------------------------------------------------------------------------- #
# Modo 1: mejorar un Excel (reemplazo de media dentro del .xlsx)
# --------------------------------------------------------------------------- #
def enhance_excel(path, out_path, max_px=1400, model=DEFAULT_MODEL):
    import hashlib
    z = zipfile.ZipFile(path)
    media = [n for n in z.namelist()
             if n.startswith("xl/media/")
             and os.path.splitext(n)[1].lstrip(".").lower() in IMG_EXTS]
    total = len(media)

    # Agrupa por contenido: imágenes idénticas (mismo artículo en varios talles)
    # se mejoran una sola vez y se reusan en todas sus copias.
    by_hash = {}            # hash -> {"arcs": [...], "raw": bytes, "ext": str}
    for n in media:
        raw = z.read(n)
        h = hashlib.md5(raw).hexdigest()
        if h not in by_hash:
            by_hash[h] = {"arcs": [], "raw": raw,
                          "ext": os.path.splitext(n)[1].lstrip(".").lower()}
        by_hash[h]["arcs"].append(n)
    uniq = len(by_hash)
    print(f"  imágenes embebidas: {total}  |  únicas a mejorar: {uniq}")

    enhanced = {}           # arcname -> bytes mejorados
    done = 0
    for h, info in by_hash.items():
        new = enhance_bytes(info["raw"], info["ext"], max_px, model=model)
        if new:
            for arc in info["arcs"]:       # aplica a todas las copias idénticas
                enhanced[arc] = new
        done += 1
        if done % 25 == 0 or done == uniq:
            print(f"    {done}/{uniq} únicas (copias cubiertas: {len(enhanced)})",
                  flush=True)
    z.close()

    # Reescribe el .xlsx cambiando solo los media mejorados
    with zipfile.ZipFile(path) as zin, \
         zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = enhanced.get(item.filename, zin.read(item.filename))
            zout.writestr(item, data)
    print(f"  ✅ {os.path.basename(out_path)}  ({len(enhanced)}/{total} en HD)")
    return len(enhanced)


def cmd_excel(args):
    if not have_engine():
        sys.exit(f"ERROR: falta el motor Real-ESRGAN en {REALESRGAN_BIN}")
    inputs = []
    for pat in args.input:
        inputs.extend(glob.glob(pat))
    inputs = [f for f in inputs
              if f.lower().endswith(".xlsx")
              and not os.path.basename(f).startswith("~$")
              and " HD.xlsx" not in f]
    if not inputs:
        sys.exit("ERROR: no encontré .xlsx para procesar.")
    for f in inputs:
        print(f"\n== {os.path.basename(f)} ==")
        out = args.output or (os.path.splitext(f)[0] + " HD.xlsx")
        if len(inputs) > 1:
            out = os.path.splitext(f)[0] + " HD.xlsx"
        enhance_excel(f, out, max_px=args.max_px, model=args.model)


# --------------------------------------------------------------------------- #
# Modo 2: descargar imágenes por artículo (alta calidad + IA opcional)
# --------------------------------------------------------------------------- #
def collect_articles(args):
    arts = list(args.articulos or [])
    if args.from_txt:
        with open(args.from_txt, encoding="utf-8") as fh:
            arts += [ln.strip() for ln in fh if ln.strip()]
    if args.from_excel:
        import openpyxl
        wb = openpyxl.load_workbook(args.from_excel, data_only=True, read_only=True)
        for ws in wb.worksheets:
            header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            want = (args.col or "Número").strip().lower()
            ci = None
            for c, name in enumerate(header, start=1):
                if name and str(name).strip().lower() == want:
                    ci = c
                    break
            if ci is None:
                continue
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, ci).value
                if v:
                    arts.append(str(v))
        wb.close()
    # dedup preservando orden
    seen, out = set(), []
    for a in arts:
        if a not in seen:
            seen.add(a)
            out.append(a)
    return out


def cmd_articulos(args):
    articles = collect_articles(args)
    if not articles:
        sys.exit("ERROR: no me pasaste artículos (args, --from-txt o --from-excel).")
    out_dir = args.output or os.path.join(BASE_DIR, "imagenes_HD")
    os.makedirs(out_dir, exist_ok=True)
    enhance = not args.no_enhance and have_engine()
    print(f"Artículos: {len(articles)} | destino: {out_dir} | IA: {enhance}")

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X "
                            "10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) "
                            "Chrome/124 Safari/537.36"})
    ok = 0
    for i, art in enumerate(articles, start=1):
        model = N.model_from_sku(art)
        url = N.reebok_image_url(model, session) if model else None
        if not url:
            continue
        try:
            r = session.get(url, timeout=20)
        except requests.RequestException:
            continue
        if r.status_code != 200 or not r.content:
            continue
        safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in str(art))
        dest = os.path.join(out_dir, f"{safe}.png")
        if enhance:
            new = enhance_bytes(r.content, "png", args.max_px, model=args.model,
                                min_px=args.max_px)  # siempre mejora lo descargado
            if new:
                with open(dest, "wb") as fh:
                    fh.write(new)
            else:
                Image.open(BytesIO(r.content)).convert("RGB").save(dest)
        else:
            Image.open(BytesIO(r.content)).convert("RGB").save(dest)
        ok += 1
        if i % 25 == 0 or i == len(articles):
            print(f"    {i}/{len(articles)} (descargadas: {ok})", flush=True)
    print(f"\n✅ {ok}/{len(articles)} imágenes en {out_dir}")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def main():
    p = argparse.ArgumentParser(description="Mejora/descarga de imágenes HD con IA")
    sub = p.add_subparsers(dest="cmd", required=True)

    pe = sub.add_parser("excel", help="Deja en HD las fotos de un .xlsx")
    pe.add_argument("--input", required=True, nargs="+",
                    help="Ruta(s) o patrón glob de .xlsx")
    pe.add_argument("--output", help="Ruta de salida (si es un solo archivo)")
    pe.add_argument("--max-px", type=int, default=1400,
                    help="Lado máximo de las imágenes mejoradas (default 1400)")
    pe.add_argument("--model", default=DEFAULT_MODEL, help="Modelo Real-ESRGAN")

    pa = sub.add_parser("articulos", help="Descarga imágenes HD por código/SKU")
    pa.add_argument("articulos", nargs="*", help="Códigos/SKU")
    pa.add_argument("--from-txt", help="Archivo .txt con un código por línea")
    pa.add_argument("--from-excel", help="Excel del que leer los códigos")
    pa.add_argument("--col", default="Número", help="Columna de código (default 'Número')")
    pa.add_argument("--output", help="Carpeta destino (default ./imagenes_HD)")
    pa.add_argument("--no-enhance", action="store_true", help="No mejorar con IA")
    pa.add_argument("--max-px", type=int, default=1400, help="Lado máximo (default 1400)")
    pa.add_argument("--model", default=DEFAULT_MODEL, help="Modelo Real-ESRGAN")

    args = p.parse_args()
    if args.cmd == "excel":
        cmd_excel(args)
    elif args.cmd == "articulos":
        cmd_articulos(args)


if __name__ == "__main__":
    main()
